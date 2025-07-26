from pathlib import Path
from typing import Any
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


def create_workbook(filename: str, sheet_name: str = "Sheet1") -> dict[str, Any]:
    """Create a new Excel workbook with optional custom sheet name"""
    try:
        wb = Workbook()
        # Rename default sheet
        if "Sheet" in wb.sheetnames:
            sheet = wb["Sheet"]
            sheet.title = sheet_name
        else:
            wb.create_sheet(sheet_name)

        path = Path(filename)
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(path))
        return {"message": f"Created workbook: {filename}", "active_sheet": sheet_name}
    except Exception as e:
        return {"error": f"Failed to create workbook: {e!s}"}


def get_or_create_workbook(filename: str) -> Workbook:
    """Get existing workbook or create new one if it doesn't exist"""
    try:
        return load_workbook(filename)
    except FileNotFoundError:
        wb = Workbook()
        wb.save(filename)
        return wb


def create_sheet(filename: str, sheet_name: str) -> dict:
    """Create a new worksheet in the workbook if it doesn't exist."""
    try:
        wb = load_workbook(filename)

        # Check if sheet already exists
        if sheet_name in wb.sheetnames:
            return {"error": f"Sheet {sheet_name} already exists"}

        # Create new sheet
        wb.create_sheet(sheet_name)
        wb.save(filename)
        wb.close()
        return {"message": f"Sheet {sheet_name} created successfully"}
    except Exception as e:
        return {"error": str(e)}


def get_workbook_info(filename: str, include_ranges: bool = False) -> dict[str, Any]:
    """Get metadata about workbook including sheets, ranges, etc."""
    try:
        path = Path(filename)
        if not path.exists():
            return {"error": f"File not found: {filename}"}
        wb = load_workbook(filename, read_only=True)
        info = {
            "filename": path.name,
            "sheets": wb.sheetnames,
            "size": path.stat().st_size,
            "modified": path.stat().st_mtime,
        }
        if include_ranges:
            ranges = {}
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                if ws.max_row > 0 and ws.max_column > 0:
                    ranges[sheet_name] = (
                        f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
                    )
            info["used_ranges"] = ranges
        wb.close()
        return info
    except Exception as e:
        return {"error": str(e)}
