import re

import psycopg2
from openpyxl import load_workbook


def fetch_data_from_db(connection_string: str, query: str) -> dict:
    try:
        conn = psycopg2.connect(connection_string)
        cur = conn.cursor()
        cur.execute(query)
        if cur.description is not None:
            rows = cur.fetchall()
            columns = [desc[0] for desc in cur.description]
        else:
            rows = []
            columns = []
        cur.close()
        conn.close()
        return {"columns": columns, "rows": rows}
    except Exception as e:
        return {"error": str(e)}


def insert_data_to_excel(
    filename: str, sheet_name: str, columns: list, rows: list
) -> dict:
    try:
        wb = load_workbook(filename)
        ws = wb[sheet_name]
        # Escribir encabezados
        for col_num, col_name in enumerate(columns, 1):
            ws.cell(row=1, column=col_num, value=col_name)
        # Escribir datos
        for row_num, row_data in enumerate(rows, 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        wb.save(filename)
        return {"message": f"Inserted {len(rows)} rows into '{sheet_name}'"}
    except Exception as e:
        return {"error": str(e)}


def validate_sql_query(query: str) -> bool:
    """
    Basic validation to prevent SQL injection in SELECT queries.
    Only allows SELECT statements without dangerous keywords.
    """
    # Only allow SELECT queries, no semicolons, no comments, no dangerous keywords
    allowed = re.match(r"^\s*SELECT\s+.+\s+FROM\s+\w+", query, re.IGNORECASE)
    forbidden = re.search(
        r"(;|--|\bDROP\b|\bDELETE\b|\bINSERT\b|\bUPDATE\b|\bALTER\b|\bTRUNCATE\b)",
        query,
        re.IGNORECASE,
    )
    return bool(allowed) and not forbidden


def clean_data(rows: list, columns: list) -> list:
    """
    Clean data rows: remove None, strip strings, and basic type normalization.
    """
    cleaned = []
    for row in rows:
        cleaned_row = []
        for value in row:
            if isinstance(value, str):
                cleaned_row.append(value.strip())
            elif value is None:
                cleaned_row.append("")
            else:
                cleaned_row.append(value)
        cleaned.append(tuple(cleaned_row))
    return cleaned


def insert_calculated_data_to_db(
    connection_string: str, table: str, columns: list, rows: list
) -> dict:
    """
    Insert calculated/cleaned data into the database.
    """
    try:
        conn = psycopg2.connect(connection_string)
        cur = conn.cursor()
        # Build parameterized query
        col_names = ",".join([f'"{col}"' for col in columns])
        placeholders = ",".join(["%s"] * len(columns))
        query = f'INSERT INTO "{table}" ({col_names}) VALUES ({placeholders})'
        for row in rows:
            cur.execute(query, row)
        conn.commit()
        cur.close()
        conn.close()
        return {"message": f"Inserted {len(rows)} rows into '{table}'"}
    except Exception as e:
        return {"error": str(e)}
