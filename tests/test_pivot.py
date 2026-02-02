"""Tests for mcp_excel.core.pivot module."""

from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook, load_workbook

from mcp_excel.core.pivot import (
    _convert_sheetdata_to_dicts,
    _get_combinations,
    _filter_data,
    _aggregate_values,
    create_pivot_table,
)
from mcp_excel.exceptions.exception_core import PivotError, ValidationError


class TestConvertSheetDataToDicts:
    """Tests for _convert_sheetdata_to_dicts function."""

    def test_convert_empty_data(self) -> None:
        """Test converting empty data."""
        result = _convert_sheetdata_to_dicts([])
        assert result == []

    def test_convert_only_headers(self) -> None:
        """Test converting data with only headers."""
        sheet_data = [["Name", "Age", "City"]]
        result = _convert_sheetdata_to_dicts(sheet_data)
        assert result == []

    def test_convert_with_data(self) -> None:
        """Test converting data with headers and rows."""
        sheet_data = [
            ["Name", "Age", "City"],
            ["Alice", 30, "NYC"],
            ["Bob", 25, "LA"],
        ]
        result = _convert_sheetdata_to_dicts(sheet_data)

        assert len(result) == 2
        assert result[0] == {"Name": "Alice", "Age": 30, "City": "NYC"}
        assert result[1] == {"Name": "Bob", "Age": 25, "City": "LA"}

    def test_convert_with_none_headers(self) -> None:
        """Test converting data with None headers."""
        sheet_data = [
            ["Name", None, "City"],
            ["Alice", 30, "NYC"],
        ]
        result = _convert_sheetdata_to_dicts(sheet_data)

        assert len(result) == 1
        assert "Name" in result[0]
        assert "Column_1" in result[0]  # None header gets default name
        assert "City" in result[0]

    def test_convert_uneven_rows(self) -> None:
        """Test converting data with uneven row lengths."""
        sheet_data = [
            ["A", "B", "C"],
            [1, 2],  # Shorter row
            [3, 4, 5, 6],  # Longer row
        ]
        result = _convert_sheetdata_to_dicts(sheet_data)

        assert len(result) == 2
        assert result[0] == {"A": 1, "B": 2}  # Missing columns are skipped
        assert result[1] == {"A": 3, "B": 4, "C": 5}  # Extra values are skipped


class TestGetCombinations:
    """Tests for _get_combinations function."""

    def test_empty_field_values(self) -> None:
        """Test with empty field values."""
        result = _get_combinations({})
        assert result == [{}]

    def test_single_field(self) -> None:
        """Test with single field."""
        field_values = {"Category": ["A", "B", "C"]}
        result = _get_combinations(field_values)

        assert len(result) == 3
        assert {"Category": "A"} in result
        assert {"Category": "B"} in result
        assert {"Category": "C"} in result

    def test_multiple_fields(self) -> None:
        """Test with multiple fields."""
        field_values = {
            "Region": ["North", "South"],
            "Product": ["X", "Y"],
        }
        result = _get_combinations(field_values)

        assert len(result) == 4
        assert {"Region": "North", "Product": "X"} in result
        assert {"Region": "North", "Product": "Y"} in result
        assert {"Region": "South", "Product": "X"} in result
        assert {"Region": "South", "Product": "Y"} in result

    def test_sorted_output(self) -> None:
        """Test that combinations are sorted."""
        field_values = {"Category": ["Z", "A", "M"]}
        result = _get_combinations(field_values)

        assert result[0]["Category"] == "A"
        assert result[1]["Category"] == "M"
        assert result[2]["Category"] == "Z"


class TestFilterData:
    """Tests for _filter_data function."""

    def test_filter_no_matches(self) -> None:
        """Test filtering with no matches."""
        data: list[dict[str, Any]] = [
            {"Region": "North", "Sales": 100},
            {"Region": "South", "Sales": 200},
        ]
        result = _filter_data(data, {"Region": "East"}, {})
        assert result == []

    def test_filter_single_match(self) -> None:
        """Test filtering with single match."""
        data: list[dict[str, Any]] = [
            {"Region": "North", "Sales": 100},
            {"Region": "South", "Sales": 200},
        ]
        result = _filter_data(data, {"Region": "North"}, {})

        assert len(result) == 1
        assert result[0]["Region"] == "North"

    def test_filter_multiple_matches(self) -> None:
        """Test filtering with multiple matches."""
        data: list[dict[str, Any]] = [
            {"Region": "North", "Product": "A", "Sales": 100},
            {"Region": "North", "Product": "B", "Sales": 200},
            {"Region": "South", "Product": "A", "Sales": 300},
        ]
        result = _filter_data(data, {"Region": "North"}, {})

        assert len(result) == 2

    def test_filter_with_col_filters(self) -> None:
        """Test filtering with column filters."""
        data: list[dict[str, Any]] = [
            {"Region": "North", "Product": "A", "Sales": 100},
            {"Region": "North", "Product": "B", "Sales": 200},
            {"Region": "South", "Product": "A", "Sales": 300},
        ]
        result = _filter_data(data, {"Region": "North"}, {"Product": "A"})

        assert len(result) == 1
        assert result[0]["Region"] == "North"
        assert result[0]["Product"] == "A"

    def test_filter_empty_data(self) -> None:
        """Test filtering empty data."""
        result = _filter_data([], {"Region": "North"}, {})
        assert result == []

    def test_filter_no_filters(self) -> None:
        """Test with no filters returns all data."""
        data: list[dict[str, Any]] = [
            {"Region": "North", "Sales": 100},
            {"Region": "South", "Sales": 200},
        ]
        result = _filter_data(data, {}, {})
        assert len(result) == 2


class TestAggregateValues:
    """Tests for _aggregate_values function."""

    def test_sum_aggregation(self) -> None:
        """Test sum aggregation."""
        data: list[dict[str, Any]] = [
            {"Sales": 100},
            {"Sales": 200},
            {"Sales": 300},
        ]
        result = _aggregate_values(data, "Sales", "sum")
        assert result == 600.0

    def test_average_aggregation(self) -> None:
        """Test average aggregation."""
        data: list[dict[str, Any]] = [
            {"Sales": 100},
            {"Sales": 200},
            {"Sales": 300},
        ]
        result = _aggregate_values(data, "Sales", "average")
        assert result == 200.0

    def test_count_aggregation(self) -> None:
        """Test count aggregation."""
        data: list[dict[str, Any]] = [
            {"Sales": 100},
            {"Sales": 200},
            {"Sales": 300},
        ]
        result = _aggregate_values(data, "Sales", "count")
        assert result == 3.0

    def test_min_aggregation(self) -> None:
        """Test min aggregation."""
        data: list[dict[str, Any]] = [
            {"Sales": 100},
            {"Sales": 200},
            {"Sales": 300},
        ]
        result = _aggregate_values(data, "Sales", "min")
        assert result == 100.0

    def test_max_aggregation(self) -> None:
        """Test max aggregation."""
        data: list[dict[str, Any]] = [
            {"Sales": 100},
            {"Sales": 200},
            {"Sales": 300},
        ]
        result = _aggregate_values(data, "Sales", "max")
        assert result == 300.0

    def test_empty_data(self) -> None:
        """Test aggregation with empty data."""
        result = _aggregate_values([], "Sales", "sum")
        assert result == 0.0

    def test_missing_field(self) -> None:
        """Test aggregation when field is missing."""
        data: list[dict[str, Any]] = [
            {"Sales": 100},
            {"Other": 200},
        ]
        result = _aggregate_values(data, "Sales", "sum")
        assert result == 100.0

    def test_non_numeric_values(self) -> None:
        """Test aggregation with non-numeric values."""
        data: list[dict[str, Any]] = [
            {"Sales": 100},
            {"Sales": "text"},
            {"Sales": None},
        ]
        result = _aggregate_values(data, "Sales", "sum")
        assert result == 100.0

    def test_default_to_sum(self) -> None:
        """Test default aggregation is sum."""
        data: list[dict[str, Any]] = [
            {"Sales": 100},
            {"Sales": 200},
        ]
        result = _aggregate_values(data, "Sales", "unknown")
        assert result == 300.0


class TestCreatePivotTable:
    """Tests for create_pivot_table function."""

    def test_create_simple_pivot_table(self, tmp_path: Path) -> None:
        """Test creating a simple pivot table."""
        test_file = tmp_path / "test_pivot.xlsx"

        # Create test data
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        # Add headers and data
        test_data = [
            ["Region", "Product", "Sales"],
            ["North", "A", 100],
            ["North", "B", 200],
            ["South", "A", 150],
            ["South", "B", 250],
        ]

        for row_idx, row_data in enumerate(test_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        wb.save(test_file)
        wb.close()

        result = create_pivot_table(
            str(test_file),
            "Data",
            "A1:C5",
            rows=["Region"],
            values=["Sales"],
            agg_func="sum",
        )

        assert "error" not in result
        assert "message" in result
        assert "Summary table created successfully" in result["message"]

        # Verify pivot sheet was created
        wb2 = load_workbook(test_file)
        assert "Data_pivot" in wb2.sheetnames
        wb2.close()

    def test_create_pivot_with_columns(self, tmp_path: Path) -> None:
        """Test creating pivot table with column fields."""
        test_file = tmp_path / "test_pivot.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        test_data = [
            ["Region", "Product", "Sales"],
            ["North", "A", 100],
            ["North", "B", 200],
            ["South", "A", 150],
            ["South", "B", 250],
        ]

        for row_idx, row_data in enumerate(test_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        wb.save(test_file)
        wb.close()

        result = create_pivot_table(
            str(test_file),
            "Data",
            "A1:C5",
            rows=["Region"],
            columns=["Product"],
            values=["Sales"],
            agg_func="sum",
        )

        assert "error" not in result
        assert "Data_pivot" in result["details"]["pivot_sheet"]

    def test_pivot_table_sheet_not_found(self, tmp_path: Path) -> None:
        """Test error when source sheet doesn't exist."""
        test_file = tmp_path / "test.xlsx"
        wb = Workbook()
        wb.active.title = "Sheet1"
        wb.save(test_file)
        wb.close()

        result = create_pivot_table(
            str(test_file),
            "NonExistent",
            "A1:B2",
            rows=["Field"],
            values=["Value"],
        )

        assert "error" in result
        assert "not found" in result["error"]

    def test_pivot_table_invalid_data_range(self, tmp_path: Path) -> None:
        """Test error with invalid data range format."""
        test_file = tmp_path / "test.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws["A1"] = "Header"
        wb.save(test_file)
        wb.close()

        result = create_pivot_table(
            str(test_file),
            "Data",
            "invalid_range",  # Missing colon
            rows=["Field"],
            values=["Value"],
        )

        assert "error" in result

    def test_pivot_table_invalid_field(self, tmp_path: Path) -> None:
        """Test error with invalid field name."""
        test_file = tmp_path / "test.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws["A1"] = "Region"
        ws["B1"] = "Sales"
        ws["A2"] = "North"
        ws["B2"] = 100
        wb.save(test_file)
        wb.close()

        result = create_pivot_table(
            str(test_file),
            "Data",
            "A1:B2",
            rows=["InvalidField"],
            values=["Sales"],
        )

        assert "error" in result
        assert "Invalid" in result["error"]

    def test_pivot_table_invalid_agg_function(self, tmp_path: Path) -> None:
        """Test error with invalid aggregation function."""
        test_file = tmp_path / "test.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws["A1"] = "Field"
        ws["A2"] = "Value"
        wb.save(test_file)
        wb.close()

        result = create_pivot_table(
            str(test_file),
            "Data",
            "A1:A2",
            rows=["Field"],
            values=["Field"],
            agg_func="invalid",
        )

        assert "error" in result
        assert "Invalid aggregation" in result["error"]

    def test_pivot_table_all_agg_functions(self, tmp_path: Path) -> None:
        """Test creating pivot table with all supported aggregation functions."""
        test_file = tmp_path / "test_pivot.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        test_data = [
            ["Region", "Sales"],
            ["North", 100],
            ["North", 200],
            ["South", 150],
            ["South", 250],
        ]

        for row_idx, row_data in enumerate(test_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        wb.save(test_file)
        wb.close()

        agg_funcs = ["sum", "average", "count", "min", "max"]

        for agg_func in agg_funcs:
            result = create_pivot_table(
                str(test_file),
                "Data",
                "A1:B5",
                rows=["Region"],
                values=["Sales"],
                agg_func=agg_func,
            )

            assert "error" not in result, f"Failed with agg_func={agg_func}"

    def test_pivot_table_overwrites_existing(self, tmp_path: Path) -> None:
        """Test that pivot table overwrites existing pivot sheet."""
        test_file = tmp_path / "test_pivot.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws["A1"] = "Field"
        ws["A2"] = "Value"

        # Pre-create a pivot sheet
        pivot_ws = wb.create_sheet("Data_pivot")
        pivot_ws["A1"] = "Old Data"

        wb.save(test_file)
        wb.close()

        result = create_pivot_table(
            str(test_file),
            "Data",
            "A1:A2",
            rows=["Field"],
            values=["Field"],
        )

        assert "error" not in result

        # Verify old data is gone
        wb2 = load_workbook(test_file)
        pivot_ws = wb2["Data_pivot"]
        # The old data should have been replaced
        wb2.close()
