"""
Test configuration and fixtures for MCP Excel Office Server tests.

This module provides pytest fixtures and configuration for testing
the MCP Excel Office Server functionality.
"""

import tempfile
from pathlib import Path
from typing import Generator
import pytest
import shutil

from mcp_excel.config import ConfigurationManager


@pytest.fixture
def test_config(tmp_path) -> Generator[ConfigurationManager, None, None]:
    """Configure the system for testing with a temporary directory."""
    manager = ConfigurationManager()

    # Store original configuration
    original_config = manager.config

    try:
        # Set up test configuration using tmp_path
        manager.reload_configuration(
            directory=str(tmp_path),
            log_level="INFO"
        )
        yield manager
    finally:
        # Restore original configuration
        manager._config = original_config


@pytest.fixture
def test_excel_file(tmp_path) -> str:
    """Create a test Excel file in the temporary directory."""
    import openpyxl

    # Configure test environment
    manager = ConfigurationManager()
    manager.reload_configuration(directory=str(tmp_path), log_level="INFO")

    # Create a test workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Add some test data
    test_data = [
        ["Name", "Age", "City"],
        ["Alice", 30, "New York"],
        ["Bob", 25, "Los Angeles"],
        ["Charlie", 35, "Chicago"]
    ]

    for row_idx, row_data in enumerate(test_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Save to tmp_path directory
    test_file_path = tmp_path / "test_file.xlsx"
    wb.save(test_file_path)

    return str(test_file_path)


@pytest.fixture
def test_workbook_file(tmp_path) -> str:
    """Create a test workbook file in the temporary directory."""
    import openpyxl

    # Configure test environment
    manager = ConfigurationManager()
    manager.reload_configuration(directory=str(tmp_path), log_level="INFO")

    # Create a test workbook with multiple sheets
    wb = openpyxl.Workbook()

    # First sheet with data
    ws1 = wb.active
    ws1.title = "Data"
    data = [
        ["Product", "Sales", "Region"],
        ["Widget A", 100, "North"],
        ["Widget B", 150, "South"],
        ["Widget C", 200, "East"]
    ]

    for row_idx, row_data in enumerate(data, 1):
        for col_idx, value in enumerate(row_data, 1):
            ws1.cell(row=row_idx, column=col_idx, value=value)

    # Second sheet for formulas
    ws2 = wb.create_sheet("Formulas")
    ws2["A1"] = "Total Sales"
    ws2["B1"] = "=SUM(Data.B2:B4)"

    # Save to tmp_path directory
    test_file_path = tmp_path / "test_workbook.xlsx"
    wb.save(test_file_path)

    return str(test_file_path)


@pytest.fixture(autouse=True)
def setup_test_documents(tmp_path) -> None:
    """Copy test documents to the test directory if they exist."""
    source_docs = Path("documents")
    if source_docs.exists():
        for doc_file in source_docs.glob("*.xlsx"):
            target_file = tmp_path / doc_file.name
            shutil.copy2(doc_file, target_file)
