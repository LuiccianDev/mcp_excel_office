"""Tests for formula security validation."""

import pytest

from mcp_excel.core.calculations import (
    apply_formula_secure,
    validate_formula_secure,
)


class TestValidateFormulaSecure:
    """Test cases for validate_formula_secure function."""

    def test_valid_simple_formula(self):
        """Test that simple valid formulas pass."""
        is_valid, message = validate_formula_secure("=SUM(A1:A10)")
        assert is_valid is True
        assert "valid" in message.lower()

    def test_valid_formula_without_equals(self):
        """Test that formulas without = prefix are validated."""
        is_valid, message = validate_formula_secure("SUM(A1:A10)")
        assert is_valid is True

    def test_valid_complex_formula(self):
        """Test that complex valid formulas pass."""
        is_valid, message = validate_formula_secure("=IF(A1>0, SUM(B1:B10), 0)")
        assert is_valid is True

    def test_valid_formula_with_text(self):
        """Test that formulas with text pass."""
        is_valid, message = validate_formula_secure('="Hello World"')
        assert is_valid is True

    def test_empty_formula_fails(self):
        """Test that empty formulas fail."""
        is_valid, message = validate_formula_secure("")
        assert is_valid is False
        assert "empty" in message.lower()

    def test_whitespace_only_formula_fails(self):
        """Test that whitespace-only formulas fail."""
        is_valid, message = validate_formula_secure("   ")
        assert is_valid is False

    def test_dangerous_function_call_fails(self):
        """Test that dangerous functions are blocked."""
        dangerous_formulas = [
            "=CALL('kernel32', 'CreateProcess', ...)",
            "=EXEC('rm -rf /')",
            "=SYSTEM('whoami')",
            "=RUN('calc.exe')",
            "=EVAL('1+1')",
        ]

        for formula in dangerous_formulas:
            is_valid, message = validate_formula_secure(formula)
            assert is_valid is False, f"Formula should fail: {formula}"
            assert "dangerous" in message.lower()

    def test_register_function_blocked(self):
        """Test that REGISTER functions are blocked."""
        is_valid, message = validate_formula_secure(
            "=REGISTER.ID('Library', 'Function')"
        )
        assert is_valid is False

    def test_xlm_functions_blocked(self):
        """Test that XLM functions are blocked."""
        is_valid, message = validate_formula_secure("=XLM.CALL(...)")
        assert is_valid is False

    def test_url_in_formula_fails(self):
        """Test that URLs in formulas are blocked."""
        formulas_with_urls = [
            "=HYPERLINK('http://evil.com')",
            "=IMPORTXML('ftp://attacker.com/file')",
            "=WEBSERVICE('https://data.leak.com')",
        ]

        for formula in formulas_with_urls:
            is_valid, message = validate_formula_secure(formula)
            # Note: HYPERLINK might be legitimate, but for security we block all URLs
            # This test might need adjustment based on actual security policy
            if "http" in formula.lower() or "ftp" in formula.lower():
                assert is_valid is False, f"Should block URL in: {formula}"

    def test_javascript_injection_blocked(self):
        """Test that JavaScript patterns are blocked."""
        formulas_with_js = [
            "=<script>alert('xss')</script>",
            "=document.write('hack')",
            "=eval('malicious_code')",
        ]

        for formula in formulas_with_js:
            is_valid, message = validate_formula_secure(formula)
            assert is_valid is False, f"Should block JS in: {formula}"

    def test_python_injection_blocked(self):
        """Test that Python code patterns are blocked."""
        formulas_with_python = [
            "=__import__('os').system('ls')",
            "=os.system('rm -rf /')",
            "=subprocess.call(['whoami'])",
        ]

        for formula in formulas_with_python:
            is_valid, message = validate_formula_secure(formula)
            assert is_valid is False, f"Should block Python in: {formula}"

    def test_unbalanced_parentheses_fails(self):
        """Test that unbalanced parentheses are detected."""
        is_valid, message = validate_formula_secure("=SUM(A1:A10")
        assert is_valid is False
        assert "parentheses" in message.lower()

    def test_extra_closing_parentheses_fails(self):
        """Test that extra closing parentheses are detected."""
        is_valid, message = validate_formula_secure("=SUM(A1:A10))")
        assert is_valid is False
        assert "parentheses" in message.lower()

    def test_formula_length_limit(self):
        """Test that formulas exceeding max length fail."""
        # Create a formula exceeding 8192 characters
        long_formula = "=" + "A1+" * 3000  # Will be > 8192 chars

        is_valid, message = validate_formula_secure(long_formula)
        assert is_valid is False
        assert "length" in message.lower()

    def test_case_insensitive_dangerous_detection(self):
        """Test that dangerous functions are detected case-insensitively."""
        formulas = [
            "=call('lib', 'func')",
            "=CALL('lib', 'func')",
            "=Call('lib', 'func')",
        ]

        for formula in formulas:
            is_valid, message = validate_formula_secure(formula)
            assert is_valid is False, f"Should detect dangerous function in: {formula}"


class TestApplyFormulaSecure:
    """Test cases for apply_formula_secure function."""

    def test_apply_formula_secure_with_valid_formula(self, tmp_path):
        """Test applying a valid formula securely."""
        from openpyxl import Workbook

        # Create test file
        test_file = tmp_path / "test_secure.xlsx"
        wb = Workbook()
        wb.save(test_file)
        wb.close()

        # Apply formula
        result = apply_formula_secure(str(test_file), "Sheet", "B1", "=SUM(A1:A5)")

        assert result["status"] == "success"

    def test_apply_formula_secure_blocks_dangerous(self, tmp_path):
        """Test that dangerous formulas are blocked."""
        from openpyxl import Workbook

        # Create test file
        test_file = tmp_path / "test_secure_block.xlsx"
        wb = Workbook()
        wb.save(test_file)
        wb.close()

        # Try to apply dangerous formula
        result = apply_formula_secure(
            str(test_file), "Sheet", "B1", "=SYSTEM('whoami')"
        )

        assert result["status"] == "error"
        assert result["security_error"] is True
        assert "Security validation failed" in result["message"]

    def test_apply_formula_secure_error_handling(self, tmp_path):
        """Test error handling in apply_formula_secure."""
        from openpyxl import Workbook

        # Create test file with Sheet1
        test_file = tmp_path / "error_test.xlsx"
        wb = Workbook()
        wb.save(test_file)
        wb.close()

        # Try with wrong sheet name
        result = apply_formula_secure(
            str(test_file), "NonExistentSheet", "A1", "=AVERAGE(1,2,3)"
        )

        # Should fail on sheet validation after passing security
        assert "status" in result
        assert result["status"] == "error"
