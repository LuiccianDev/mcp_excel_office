"""Tests for mcp_excel.core.data module."""

from pathlib import Path

import pytest
from openpyxl import Workbook

from mcp_excel.core.data import (
    _get_used_range,
    _get_worksheet,
    _parse_cell_reference,
    _write_data_to_worksheet,
    read_excel_range,
    write_data,
)
from mcp_excel.exceptions.exception_core import (
    DataError,
    InvalidCellReferenceError,
    InvalidDataError,
    RangeError,
    SheetNotFoundError,
    WorkbookError,
)


class TestGetWorksheet:
    """Tests for _get_worksheet function."""

    def test_get_existing_worksheet(self, tmp_path: Path) -> None:
        """Test getting an existing worksheet."""
        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"

        result = _get_worksheet(wb, "TestSheet")

        assert result.title == "TestSheet"
        wb.close()

    def test_get_nonexistent_worksheet(self, tmp_path: Path) -> None:
        """Test getting a non-existent worksheet raises error."""
        wb = Workbook()

        with pytest.raises(SheetNotFoundError) as exc_info:
            _get_worksheet(wb, "NonExistent")

        assert "NonExistent" in str(exc_info.value)
        wb.close()


class TestParseCellReference:
    """Tests for _parse_cell_reference function."""

    def test_valid_cell_a1(self) -> None:
        """Test parsing valid cell reference A1."""
        result = _parse_cell_reference("A1")
        assert result == (1, 1)

    def test_valid_cell_b2(self) -> None:
        """Test parsing valid cell reference B2."""
        result = _parse_cell_reference("B2")
        assert result == (2, 2)

    def test_valid_cell_aa10(self) -> None:
        """Test parsing valid cell reference AA10."""
        result = _parse_cell_reference("AA10")
        assert result == (10, 27)

    def test_invalid_cell_reference(self) -> None:
        """Test parsing invalid cell reference raises error."""
        with pytest.raises(InvalidCellReferenceError) as exc_info:
            _parse_cell_reference("invalid")

        assert "Invalid cell reference" in str(exc_info.value)

    def test_empty_cell_reference(self) -> None:
        """Test parsing empty cell reference raises error."""
        with pytest.raises(InvalidCellReferenceError):
            _parse_cell_reference("")


class TestGetUsedRange:
    """Tests for _get_used_range function."""

    def test_get_used_range_with_data(self, tmp_path: Path) -> None:
        """Test getting used range with data."""
        wb = Workbook()
        ws = wb.active

        # Add data to A1:C3
        ws["A1"] = "Header1"
        ws["B1"] = "Header2"
        ws["C1"] = "Header3"
        ws["A2"] = "Data1"
        ws["B2"] = "Data2"
        ws["C2"] = "Data3"

        end_row, end_col = _get_used_range(ws, 1, 1)

        # Should detect the used range starting from A1
        assert end_row >= 2
        assert end_col >= 3
        wb.close()

    def test_get_used_range_empty(self, tmp_path: Path) -> None:
        """Test getting used range on empty sheet."""
        wb = Workbook()
        ws = wb.active

        end_row, end_col = _get_used_range(ws, 1, 1)

        # Should return starting coordinates for empty sheet
        assert end_row == 0
        assert end_col == 0
        wb.close()


class TestReadExcelRange:
    """Tests for read_excel_range function."""

    def test_read_nonexistent_file(self, tmp_path: Path) -> None:
        """Test reading from non-existent file raises error."""
        nonexistent_file = tmp_path / "nonexistent.xlsx"

        with pytest.raises(FileNotFoundError):
            read_excel_range(nonexistent_file, "Sheet1")

    def test_read_nonexistent_sheet(self, tmp_path: Path) -> None:
        """Test reading from non-existent sheet raises error."""
        test_file = tmp_path / "test.xlsx"
        wb = Workbook()
        wb.active.title = "Sheet1"
        wb.save(test_file)
        wb.close()

        with pytest.raises(SheetNotFoundError) as exc_info:
            read_excel_range(test_file, "NonExistent")

        assert "NonExistent" in str(exc_info.value)

    def test_read_valid_range(self, tmp_path: Path) -> None:
        """Test reading a valid range."""
        test_file = tmp_path / "test.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        test_data = [
            ["Name", "Age", "City"],
            ["Alice", 30, "NYC"],
            ["Bob", 25, "LA"],
        ]

        for row_idx, row_data in enumerate(test_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        wb.save(test_file)
        wb.close()

        result = read_excel_range(test_file, "Sheet1", "A1", "C3")

        assert len(result) == 3
        assert result[0] == ["Name", "Age", "City"]
        assert result[1] == ["Alice", 30, "NYC"]

    def test_read_range_with_auto_detection(self, tmp_path: Path) -> None:
        """Test reading range with automatic end cell detection."""
        test_file = tmp_path / "test.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        ws["A1"] = "Header"
        ws["A2"] = "Data1"
        ws["A3"] = "Data2"

        wb.save(test_file)
        wb.close()

        result = read_excel_range(test_file, "Sheet1", "A1")

        assert len(result) >= 1

    def test_read_out_of_bounds_range(self, tmp_path: Path) -> None:
        """Test reading out of bounds range raises error."""
        test_file = tmp_path / "test.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "Data"
        wb.save(test_file)
        wb.close()

        with pytest.raises(RangeError) as exc_info:
            read_excel_range(test_file, "Sheet1", "Z100")

        assert "out of bounds" in str(exc_info.value)


class TestWriteDataToWorksheet:
    """Tests for _write_data_to_worksheet function."""

    def test_write_data_to_worksheet(self, tmp_path: Path) -> None:
        """Test writing data to worksheet."""
        wb = Workbook()
        ws = wb.active

        test_data = [
            ["Name", "Age"],
            ["Alice", 30],
            ["Bob", 25],
        ]

        _write_data_to_worksheet(ws, test_data, 1, 1)

        assert ws["A1"].value == "Name"
        assert ws["B1"].value == "Age"
        assert ws["A2"].value == "Alice"
        assert ws["B2"].value == 30
        wb.close()

    def test_write_data_with_offset(self, tmp_path: Path) -> None:
        """Test writing data with row and column offset."""
        wb = Workbook()
        ws = wb.active

        test_data = [["Data1", "Data2"]]

        _write_data_to_worksheet(ws, test_data, 5, 3)

        assert ws["C5"].value == "Data1"
        assert ws["D5"].value == "Data2"
        wb.close()

    def test_write_data_preserves_none_values(self, tmp_path: Path) -> None:
        """Test that None values are skipped to preserve formatting."""
        wb = Workbook()
        ws = wb.active

        # Pre-set a value
        ws["A1"].value = "Existing"

        test_data = [[None, "New"]]

        _write_data_to_worksheet(ws, test_data, 1, 1)

        # None should not overwrite existing value
        assert ws["A1"].value == "Existing"
        assert ws["B1"].value == "New"
        wb.close()


class TestWriteData:
    """Tests for write_data function."""

    def test_write_data_new_file(self, tmp_path: Path) -> None:
        """Test writing data to new file."""
        test_file = tmp_path / "new_file.xlsx"

        test_data = [
            ["Name", "Value"],
            ["Item1", 100],
        ]

        result = write_data(test_file, "Sheet1", test_data)

        assert result["status"] == "success"
        assert "Sheet1" in result["message"]
        assert test_file.exists()

    def test_write_data_existing_file(self, tmp_path: Path) -> None:
        """Test writing data to existing file."""
        test_file = tmp_path / "existing.xlsx"
        wb = Workbook()
        wb.active.title = "Sheet1"
        wb.save(test_file)
        wb.close()

        test_data = [["New", "Data"]]

        result = write_data(test_file, "Sheet1", test_data)

        assert result["status"] == "success"

    def test_write_data_create_new_sheet(self, tmp_path: Path) -> None:
        """Test creating new sheet when writing data."""
        test_file = tmp_path / "test.xlsx"
        wb = Workbook()
        wb.active.title = "Sheet1"
        wb.save(test_file)
        wb.close()

        test_data = [["Data"]]

        result = write_data(test_file, "NewSheet", test_data)

        assert result["status"] == "success"
        assert "NewSheet" in result["active_sheet"]

    def test_write_data_no_data_raises_error(self, tmp_path: Path) -> None:
        """Test that writing empty data raises error."""
        test_file = tmp_path / "test.xlsx"

        with pytest.raises(InvalidDataError) as exc_info:
            write_data(test_file, "Sheet1", None)

        assert "No data provided" in str(exc_info.value)

    def test_write_data_empty_data_raises_error(self, tmp_path: Path) -> None:
        """Test that writing empty list raises error."""
        test_file = tmp_path / "test.xlsx"

        with pytest.raises(InvalidDataError):
            write_data(test_file, "Sheet1", [])

    def test_write_data_with_custom_start_cell(self, tmp_path: Path) -> None:
        """Test writing data with custom start cell."""
        test_file = tmp_path / "test.xlsx"

        test_data = [["A", "B"], ["C", "D"]]

        result = write_data(test_file, "Sheet1", test_data, "C3")

        assert result["status"] == "success"

        # Verify file contents
        wb = Workbook()
        # Note: write_data creates the file, we need to reload it
        from openpyxl import load_workbook

        wb2 = load_workbook(test_file)
        ws = wb2.active
        assert ws["C3"].value == "A"
        assert ws["D3"].value == "B"
        assert ws["C4"].value == "C"
        wb2.close()

    def test_write_data_invalid_cell_reference(self, tmp_path: Path) -> None:
        """Test writing data with invalid cell reference raises error."""
        test_file = tmp_path / "test.xlsx"

        test_data = [["Data"]]

        with pytest.raises(InvalidCellReferenceError):
            write_data(test_file, "Sheet1", test_data, "invalid")
