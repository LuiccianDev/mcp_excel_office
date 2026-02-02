"""Tests for mcp_excel.core.workbook module."""

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from mcp_excel.core.workbook import (
    create_workbook,
    create_sheet,
    get_or_create_workbook,
    get_workbook_info,
    managed_workbook,
    managed_worksheet,
    _create_new_sheet,
    _create_initial_worksheet,
    _load_existing_workbook,
    _validate_sheet_name,
    _get_worksheet_range,
)
from mcp_excel.exceptions.exception_core import (
    SheetExistsError,
    SheetNotFoundError,
    ValidationError,
    WorkbookError,
)


class TestCreateWorkbook:
    """Tests for create_workbook function."""

    def test_create_new_workbook(self, tmp_path: Path) -> None:
        """Test creating a new workbook."""
        test_file = tmp_path / "new_workbook.xlsx"

        result = create_workbook(test_file, "TestSheet")

        assert result["status"] == "success"
        assert "Created workbook" in result["message"]
        assert result["active_sheet"] == "TestSheet"
        assert test_file.exists()

    def test_create_workbook_default_sheet_name(self, tmp_path: Path) -> None:
        """Test creating workbook with default sheet name."""
        test_file = tmp_path / "default.xlsx"

        result = create_workbook(test_file)

        assert result["active_sheet"] == "Sheet1"

    def test_create_workbook_invalid_sheet_name(self, tmp_path: Path) -> None:
        """Test creating workbook with invalid sheet name raises error."""
        test_file = tmp_path / "test.xlsx"

        with pytest.raises((ValidationError, WorkbookError)) as exc_info:
            create_workbook(test_file, "Invalid/Name")

        error_msg = str(exc_info.value)
        assert "character" in error_msg or "cannot contain" in error_msg or "Failed to create" in error_msg


class TestGetOrCreateWorkbook:
    """Tests for get_or_create_workbook function."""

    def test_get_existing_workbook(self, tmp_path: Path) -> None:
        """Test getting an existing workbook."""
        test_file = tmp_path / "existing.xlsx"
        wb = Workbook()
        wb.active.title = "Data"
        wb.save(test_file)
        wb.close()

        result = get_or_create_workbook(test_file)

        assert result is not None
        assert "Data" in result.sheetnames
        result.close()

    def test_create_new_workbook_if_not_exists(self, tmp_path: Path) -> None:
        """Test creating new workbook when file doesn't exist."""
        test_file = tmp_path / "new.xlsx"

        result = get_or_create_workbook(test_file)

        assert result is not None
        assert "Sheet1" in result.sheetnames
        result.close()

    def test_get_existing_read_only(self, tmp_path: Path) -> None:
        """Test getting existing workbook in read-only mode."""
        test_file = tmp_path / "readonly.xlsx"
        wb = Workbook()
        wb.save(test_file)
        wb.close()

        result = get_or_create_workbook(test_file, read_only=True)

        assert result is not None
        result.close()

    def test_read_only_nonexistent_raises_error(self, tmp_path: Path) -> None:
        """Test that read-only mode with non-existent file raises error."""
        test_file = tmp_path / "nonexistent.xlsx"

        with pytest.raises(FileNotFoundError):
            get_or_create_workbook(test_file, read_only=True)


class TestCreateSheet:
    """Tests for create_sheet function."""

    def test_create_sheet_in_existing_workbook(self, tmp_path: Path) -> None:
        """Test creating a new sheet in existing workbook."""
        test_file = tmp_path / "test.xlsx"
        wb = Workbook()
        wb.active.title = "Sheet1"
        wb.save(test_file)
        wb.close()

        result = create_sheet(test_file, "NewSheet")

        assert result["status"] == "success"
        assert "NewSheet" in result["message"]

        # Verify sheet was created
        wb2 = load_workbook(test_file)
        assert "NewSheet" in wb2.sheetnames
        wb2.close()

    def test_create_sheet_in_new_workbook(self, tmp_path: Path) -> None:
        """Test creating sheet creates new workbook if file doesn't exist."""
        test_file = tmp_path / "new.xlsx"

        result = create_sheet(test_file, "FirstSheet")

        assert result["status"] == "success"
        assert test_file.exists()

    def test_create_duplicate_sheet_raises_error(self, tmp_path: Path) -> None:
        """Test creating duplicate sheet raises error."""
        test_file = tmp_path / "test.xlsx"
        wb = Workbook()
        wb.active.title = "Existing"
        wb.save(test_file)
        wb.close()

        with pytest.raises(SheetExistsError) as exc_info:
            create_sheet(test_file, "Existing")

        assert "already exists" in str(exc_info.value)

    def test_create_sheet_invalid_name(self, tmp_path: Path) -> None:
        """Test creating sheet with invalid name raises error."""
        test_file = tmp_path / "test.xlsx"
        wb = Workbook()
        wb.save(test_file)
        wb.close()

        with pytest.raises(ValidationError):
            create_sheet(test_file, "Name*Invalid")


class TestGetWorkbookInfo:
    """Tests for get_workbook_info function."""

    def test_get_basic_workbook_info(self, tmp_path: Path) -> None:
        """Test getting basic workbook info."""
        test_file = tmp_path / "test.xlsx"
        wb = Workbook()
        wb.active.title = "Data"
        wb.create_sheet("Summary")
        wb.save(test_file)
        wb.close()

        result = get_workbook_info(test_file)

        assert result["filename"] == "test.xlsx"
        assert "Data" in result["sheets"]
        assert "Summary" in result["sheets"]
        assert result["size"] > 0
        assert result["modified"] > 0

    def test_get_workbook_info_with_ranges(self, tmp_path: Path) -> None:
        """Test getting workbook info with used ranges."""
        test_file = tmp_path / "test.xlsx"
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Data"
        ws["B2"] = "More"
        wb.save(test_file)
        wb.close()

        result = get_workbook_info(test_file, include_ranges=True)

        assert result["used_ranges"] is not None
        assert "A1" in str(result["used_ranges"])

    def test_get_workbook_info_nonexistent_file(self, tmp_path: Path) -> None:
        """Test getting info for non-existent file raises error."""
        test_file = tmp_path / "nonexistent.xlsx"

        with pytest.raises(FileNotFoundError):
            get_workbook_info(test_file)


class TestValidateSheetName:
    """Tests for _validate_sheet_name function."""

    def test_valid_sheet_name(self) -> None:
        """Test validating valid sheet names."""
        # Should not raise
        _validate_sheet_name("ValidName")
        _validate_sheet_name("Sheet1")
        _validate_sheet_name("A_B_C")

    def test_empty_sheet_name(self) -> None:
        """Test empty sheet name raises error."""
        with pytest.raises(ValidationError) as exc_info:
            _validate_sheet_name("")

        assert "non-empty" in str(exc_info.value)

    def test_none_sheet_name(self) -> None:
        """Test None sheet name raises error."""
        with pytest.raises(ValidationError):
            _validate_sheet_name(None)

    def test_long_sheet_name(self) -> None:
        """Test sheet name exceeding 31 characters raises error."""
        with pytest.raises(ValidationError) as exc_info:
            _validate_sheet_name("A" * 32)

        assert "31" in str(exc_info.value)

    def test_invalid_characters(self) -> None:
        """Test sheet name with invalid characters raises error."""
        invalid_names = ["Sheet/Name", "Sheet\\Name", "Sheet*Name", "Sheet?Name", "Sheet:Name"]

        for name in invalid_names:
            with pytest.raises(ValidationError):
                _validate_sheet_name(name)

    def test_starting_with_quote(self) -> None:
        """Test sheet name starting with quote raises error."""
        with pytest.raises(ValidationError):
            _validate_sheet_name("'Sheet")


class TestCreateInitialWorksheet:
    """Tests for _create_initial_worksheet function."""

    def test_create_initial_worksheet(self) -> None:
        """Test creating initial worksheet."""
        wb = Workbook()

        _create_initial_worksheet(wb, "MySheet")

        assert "MySheet" in wb.sheetnames
        wb.close()


class TestLoadExistingWorkbook:
    """Tests for _load_existing_workbook function."""

    def test_load_existing_file(self, tmp_path: Path) -> None:
        """Test loading existing workbook."""
        test_file = tmp_path / "test.xlsx"
        wb = Workbook()
        wb.active.title = "Data"
        wb.save(test_file)
        wb.close()

        result = _load_existing_workbook(test_file)

        assert result is not None
        assert "Data" in result.sheetnames
        result.close()

    def test_load_nonexistent_file(self, tmp_path: Path) -> None:
        """Test loading non-existent file raises error."""
        test_file = tmp_path / "nonexistent.xlsx"

        with pytest.raises(WorkbookError):
            _load_existing_workbook(test_file)


class TestCreateNewSheet:
    """Tests for _create_new_sheet function."""

    def test_create_new_sheet_success(self) -> None:
        """Test creating a new sheet successfully."""
        wb = Workbook()

        _create_new_sheet(wb, "NewSheet")

        assert "NewSheet" in wb.sheetnames
        wb.close()

    def test_create_duplicate_sheet_raises_error(self) -> None:
        """Test creating duplicate sheet raises error."""
        wb = Workbook()
        wb.active.title = "Existing"

        with pytest.raises(SheetExistsError):
            _create_new_sheet(wb, "Existing")

        wb.close()

    def test_create_sheet_invalid_workbook_type(self) -> None:
        """Test creating sheet with invalid workbook type raises error."""
        with pytest.raises(TypeError):
            _create_new_sheet("not_a_workbook", "Sheet")


class TestGetWorksheetRange:
    """Tests for _get_worksheet_range function."""

    def test_get_range_with_data(self) -> None:
        """Test getting range from worksheet with data."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Data"
        ws["C3"] = "More"

        result = _get_worksheet_range(ws)

        assert result is not None
        assert "A1" in result
        assert "C3" in result
        wb.close()

    def test_get_range_empty_sheet(self) -> None:
        """Test getting range from empty worksheet."""
        wb = Workbook()
        ws = wb.active

        result = _get_worksheet_range(ws)

        # Empty sheet might still return a range depending on openpyxl version
        wb.close()


class TestManagedWorkbook:
    """Tests for managed_workbook context manager."""

    def test_managed_workbook_creates_file(self, tmp_path: Path) -> None:
        """Test managed workbook creates and saves file."""
        test_file = tmp_path / "managed.xlsx"

        with managed_workbook(test_file) as wb:
            ws = wb.active
            ws["A1"] = "Managed Data"

        assert test_file.exists()

        # Verify data was saved
        wb2 = load_workbook(test_file)
        assert wb2.active["A1"].value == "Managed Data"
        wb2.close()

    def test_managed_workbook_read_only(self, tmp_path: Path) -> None:
        """Test managed workbook in read-only mode."""
        test_file = tmp_path / "readonly.xlsx"
        wb = Workbook()
        wb.active["A1"] = "Existing"
        wb.save(test_file)
        wb.close()

        with managed_workbook(test_file, read_only=True) as wb:
            # Should be able to read
            assert wb.active["A1"].value == "Existing"


class TestManagedWorksheet:
    """Tests for managed_worksheet context manager."""

    def test_managed_worksheet_existing(self, tmp_path: Path) -> None:
        """Test accessing existing worksheet."""
        test_file = tmp_path / "test.xlsx"
        wb = Workbook()
        wb.active.title = "Data"
        wb.save(test_file)
        wb.close()

        with managed_worksheet(test_file, "Data") as ws:
            ws["A1"] = "Test Value"

        # Verify changes were saved
        wb2 = load_workbook(test_file)
        assert wb2["Data"]["A1"].value == "Test Value"
        wb2.close()

    def test_managed_worksheet_create_if_missing(self, tmp_path: Path) -> None:
        """Test creating worksheet if it doesn't exist."""
        test_file = tmp_path / "test.xlsx"
        wb = Workbook()
        wb.save(test_file)
        wb.close()

        with managed_worksheet(test_file, "NewSheet", create_if_missing=True) as ws:
            ws["A1"] = "Created"

        wb2 = load_workbook(test_file)
        assert "NewSheet" in wb2.sheetnames
        wb2.close()

    def test_managed_worksheet_not_found_raises_error(self, tmp_path: Path) -> None:
        """Test that non-existent worksheet raises error when create_if_missing is False."""
        test_file = tmp_path / "test.xlsx"
        wb = Workbook()
        wb.save(test_file)
        wb.close()

        with pytest.raises(SheetNotFoundError):
            with managed_worksheet(test_file, "Missing", create_if_missing=False):
                pass
