from pathlib import Path
from unittest.mock import patch

import pytest

from openpyxl import Workbook

from mcp_excel.exceptions.exception_tools import ValidationError
from mcp_excel.tools import formulas_excel_tools
from mcp_excel.config import ConfigurationManager
from mcp_excel.utils.validation_utils import validate_formula_in_cell_operation


# Test data
TEST_SHEET = "Sheet1"
TEST_CELL = "A1"
TEST_FORMULA = "=SUM(A1:A10)"
TEST_INVALID_FORMULA = "=SUM(A1:A10"  # Missing closing parenthesis


def create_test_workbook(tmp_path: Path, filename: str = "test_workbook.xlsx") -> str:
    """Create a test workbook with a sheet."""
    test_file = str(tmp_path / filename)
    wb = Workbook()
    wb.active.title = TEST_SHEET
    wb.save(test_file)
    return test_file


# Tests for validate_formula_in_cell_operation (real implementation)
def test_validate_formula_in_cell_operation_valid(tmp_path: Path) -> None:
    """Test real validation with valid formula."""
    test_file = create_test_workbook(tmp_path)

    result = validate_formula_in_cell_operation(
        filepath=test_file,
        sheet_name=TEST_SHEET,
        cell=TEST_CELL,
        formula=TEST_FORMULA,
    )

    assert result["status"] == "success"
    assert result["valid"] is True
    assert result["cell"] == TEST_CELL
    assert result["formula"] == TEST_FORMULA
    assert "ready to be applied" in result["message"]


def test_validate_formula_in_cell_operation_invalid_syntax(tmp_path: Path) -> None:
    """Test real validation with invalid formula syntax."""
    test_file = create_test_workbook(tmp_path)

    result = validate_formula_in_cell_operation(
        filepath=test_file,
        sheet_name=TEST_SHEET,
        cell=TEST_CELL,
        formula=TEST_INVALID_FORMULA,
    )

    assert result["status"] == "error"
    assert "Invalid formula syntax" in result["message"]


def test_validate_formula_in_cell_operation_invalid_cell(tmp_path: Path) -> None:
    """Test real validation with invalid cell reference."""
    test_file = create_test_workbook(tmp_path)

    result = validate_formula_in_cell_operation(
        filepath=test_file,
        sheet_name=TEST_SHEET,
        cell="INVALID",
        formula=TEST_FORMULA,
    )

    assert result["status"] == "error"
    assert "Invalid cell reference" in result["message"]


def test_validate_formula_in_cell_operation_sheet_not_found(tmp_path: Path) -> None:
    """Test real validation with non-existent sheet."""
    test_file = create_test_workbook(tmp_path)

    result = validate_formula_in_cell_operation(
        filepath=test_file,
        sheet_name="NonExistentSheet",
        cell=TEST_CELL,
        formula=TEST_FORMULA,
    )

    assert result["status"] == "error"
    assert "Sheet 'NonExistentSheet' not found" in result["message"]


# Tests for validate_formula_syntax (async wrapper)
@pytest.mark.asyncio  # type: ignore[misc]
async def test_validate_formula_syntax_valid(tmp_path: Path) -> None:
    """Test successful formula validation via async wrapper."""
    # Configure test environment
    manager = ConfigurationManager()
    manager.reload_configuration(directory=str(tmp_path), log_level="INFO")

    test_file = create_test_workbook(tmp_path)

    result = await formulas_excel_tools.validate_formula_syntax(
        filename=test_file,
        sheet_name=TEST_SHEET,
        cell=TEST_CELL,
        formula=TEST_FORMULA,
    )

    assert result["status"] == "success"
    assert result["valid"] is True
    assert result["formula"] == TEST_FORMULA
    assert result["cell"] == TEST_CELL


@pytest.mark.asyncio  # type: ignore[misc]
async def test_validate_formula_syntax_invalid(tmp_path: Path) -> None:
    """Test validation of invalid formula via async wrapper."""
    # Configure test environment
    manager = ConfigurationManager()
    manager.reload_configuration(directory=str(tmp_path), log_level="INFO")

    test_file = create_test_workbook(tmp_path)

    result = await formulas_excel_tools.validate_formula_syntax(
        filename=test_file,
        sheet_name=TEST_SHEET,
        cell=TEST_CELL,
        formula=TEST_INVALID_FORMULA,
    )

    assert result["status"] == "error"
    assert "Invalid formula syntax" in result["message"]


# Tests for apply_formula_excel
@patch(
    "mcp_excel.tools.formulas_excel_tools.validate_file_access",
    lambda arg: (lambda f: f),
)
@pytest.mark.asyncio  # type: ignore[misc]
async def test_apply_formula_success(tmp_path: Path) -> None:
    """Test successful formula application."""
    # Configure test environment
    manager = ConfigurationManager()
    manager.reload_configuration(directory=str(tmp_path), log_level="INFO")

    test_file = create_test_workbook(tmp_path)

    with patch("mcp_excel.tools.formulas_excel_tools.apply_formula") as mock_apply:
        # Simulate successful application
        mock_apply.return_value = {
            "status": "success",
            "cell": TEST_CELL,
            "formula": TEST_FORMULA,
            "message": "Applied formula successfully",
        }

        result = await formulas_excel_tools.apply_formula_excel(
            filename=test_file,
            sheet_name=TEST_SHEET,
            cell=TEST_CELL,
            formula=TEST_FORMULA,
        )

        # Assertions
        assert result["status"] == "success"
        assert result["cell"] == TEST_CELL
        assert result["formula"] == TEST_FORMULA

        # Verify apply_formula was called
        mock_apply.assert_called_once_with(
            test_file, TEST_SHEET, TEST_CELL, TEST_FORMULA
        )


@patch(
    "mcp_excel.tools.formulas_excel_tools.validate_file_access",
    lambda arg: (lambda f: f),
)
@pytest.mark.asyncio  # type: ignore[misc]
async def test_apply_formula_validation_failure(tmp_path: Path) -> None:
    """Test formula application with invalid formula."""
    # Configure test environment
    manager = ConfigurationManager()
    manager.reload_configuration(directory=str(tmp_path), log_level="INFO")

    test_file = create_test_workbook(tmp_path)

    result = await formulas_excel_tools.apply_formula_excel(
        filename=test_file,
        sheet_name=TEST_SHEET,
        cell=TEST_CELL,
        formula=TEST_INVALID_FORMULA,
    )

    assert result["status"] == "error"
    assert "Invalid formula syntax" in result["message"]


@pytest.mark.asyncio  # type: ignore[misc]
async def test_apply_formula_real_integration(tmp_path: Path) -> None:
    """Integration test: apply formula to real workbook."""
    # Configure test environment
    manager = ConfigurationManager()
    manager.reload_configuration(directory=str(tmp_path), log_level="INFO")

    test_file = create_test_workbook(tmp_path)

    # Add some data to the workbook
    from openpyxl import load_workbook

    wb = load_workbook(test_file)
    ws = wb[TEST_SHEET]
    ws["A1"] = 10
    ws["A2"] = 20
    wb.save(test_file)

    # Apply formula
    result = await formulas_excel_tools.apply_formula_excel(
        filename=test_file,
        sheet_name=TEST_SHEET,
        cell="A3",
        formula="=A1+A2",
    )

    assert result["status"] == "success"
    assert result["cell"] == "A3"
    assert result["formula"] == "=A1+A2"

    # Verify the formula was actually saved
    wb = load_workbook(test_file, data_only=False)
    ws = wb[TEST_SHEET]
    assert ws["A3"].value == "=A1+A2"
