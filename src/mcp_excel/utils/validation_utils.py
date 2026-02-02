import re
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from mcp_excel.utils.cell_utils import parse_cell_range, validate_cell_reference


def validate_formula_in_cell_operation(
    filepath: str, sheet_name: str, cell: str, formula: str
) -> dict[str, Any]:
    # Quick validations
    if not validate_cell_reference(cell):
        return {"status": "error", "message": f"Invalid cell reference: {cell}"}
    try:
        wb = load_workbook(filepath)
        if sheet_name not in wb.sheetnames:
            return {"status": "error", "message": f"Sheet '{sheet_name}' not found"}
        # Validate formula syntax
        result: tuple[bool, str] = validate_formula(formula)
        is_valid, message = result
        if not is_valid:
            return {"status": "error", "message": f"Invalid formula syntax: {message}"}
        # Validate cell references in the formula
        cell_refs = re.findall(r"[A-Z]+[0-9]+(?::[A-Z]+[0-9]+)?", formula)
        for ref in cell_refs:
            if ":" in ref:
                start, end = ref.split(":")
                if not (
                    validate_cell_reference(start) and validate_cell_reference(end)
                ):
                    return {
                        "status": "error",
                        "message": f"Invalid cell range reference in formula: {ref}",
                    }
            else:
                if not validate_cell_reference(ref):
                    return {
                        "status": "error",
                        "message": f"Invalid cell reference in formula: {ref}",
                    }

        # All validations passed - formula is valid and can be applied
        return {
            "status": "success",
            "message": f"Formula '{formula}' is valid and ready to be applied to cell {cell}",
            "valid": True,
            "cell": cell,
            "formula": formula,
        }
    except Exception as e:
        return {"status": "error", "message": str(e)}


def validate_range_in_sheet_operation(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: str | None = None,
) -> dict[str, Any]:
    if not validate_cell_reference(start_cell):
        return {
            "status": "error",
            "message": f"Invalid start cell reference: {start_cell}",
        }
    if end_cell and not validate_cell_reference(end_cell):
        return {"status": "error", "message": f"Invalid end cell reference: {end_cell}"}
    try:
        wb = load_workbook(filepath)
        if sheet_name not in wb.sheetnames:
            return {"status": "error", "message": f"Sheet '{sheet_name}' not found"}
        worksheet = wb[sheet_name]
        data_max_row = worksheet.max_row
        data_max_col = worksheet.max_column
        try:
            start_row, start_col, end_row, end_col = parse_cell_range(
                start_cell, end_cell
            )
        except ValueError as e:
            return {"status": "error", "message": f"Invalid range: {str(e)}"}
        if end_row is None:
            end_row = start_row
        if end_col is None:
            end_col = start_col
        is_valid, message = validate_range_bounds(
            worksheet, start_row, start_col, end_row, end_col
        )
        if not is_valid:
            return {"status": "error", "message": message}
        range_str = f"{start_cell}" if end_cell is None else f"{start_cell}:{end_cell}"
        data_range_str = f"A1:{get_column_letter(data_max_col)}{data_max_row}"
        extends_beyond_data = end_row > data_max_row or end_col > data_max_col
        return {
            "status": "success",
            "message": (
                f"Range '{range_str}' is valid. "
                f"Sheet contains data in range '{data_range_str}'"
            ),
            "valid": True,
            "range": range_str,
            "data_range": data_range_str,
            "extends_beyond_data": extends_beyond_data,
            "data_dimensions": {
                "max_row": data_max_row,
                "max_col": data_max_col,
                "max_col_letter": get_column_letter(data_max_col),
            },
        }
    except Exception as e:
        return {"status": "error", "message": str(e)}


def validate_formula(formula: str) -> tuple[bool, str]:
    if not formula.startswith("="):
        return False, "Formula must start with '='"
    formula = formula[1:]
    parens = 0
    for c in formula:
        if c == "(":
            parens += 1
        elif c == ")":
            parens -= 1
        if parens < 0:
            return False, "Unmatched closing parenthesis"
    if parens > 0:
        return False, "Unclosed parenthesis"
    func_pattern = r"([A-Z]+)\("
    funcs = re.findall(func_pattern, formula)
    unsafe_funcs = {"INDIRECT", "HYPERLINK", "WEBSERVICE", "DGET", "RTD"}
    for func in funcs:
        if func in unsafe_funcs:
            return False, f"Unsafe function: {func}"
    return True, "Formula is valid"


def validate_range_bounds(
    worksheet: Worksheet,
    start_row: int,
    start_col: int,
    end_row: int | None = None,
    end_col: int | None = None,
) -> tuple[bool, str]:
    max_row = worksheet.max_row
    max_col = worksheet.max_column
    try:
        if start_row < 1 or start_row > max_row:
            return False, f"Start row {start_row} out of bounds (1-{max_row})"
        if start_col < 1 or start_col > max_col:
            return False, (
                f"Start column {get_column_letter(start_col)} "
                f"out of bounds (A-{get_column_letter(max_col)})"
            )
        if end_row is not None and end_col is not None:
            if end_row < start_row:
                return False, "End row cannot be before start row"
            if end_col < start_col:
                return False, "End column cannot be before start column"
            if end_row > max_row:
                return False, f"End row {end_row} out of bounds (1-{max_row})"
            if end_col > max_col:
                return False, (
                    f"End column {get_column_letter(end_col)} "
                    f"out of bounds (A-{get_column_letter(max_col)})"
                )
        return True, "Range is valid"
    except Exception as e:
        return False, f"Invalid range: {e!s}"
