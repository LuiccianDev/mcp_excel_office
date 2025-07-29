from typing import Any, Literal, TypedDict

from openpyxl.formatting.rule import (
    CellIsRule,
    ColorScaleRule,
    DataBarRule,
    FormulaRule,
    IconSetRule,
)
from openpyxl.styles import (
    Alignment,
    Border,
    Color,
    Font,
    PatternFill,
    Protection,
    Side,
)

from mcp_excel.core.workbook import get_or_create_workbook
from mcp_excel.utils.cell_utils import parse_cell_range, validate_cell_reference

# Define valid border styles
BorderStyle = Literal[
    'dashDot',
    'dashDotDot',
    'dashed',
    'dotted',
    'double',
    'hair',
    'medium',
    'mediumDashDot',
    'mediumDashDotDot',
    'mediumDashed',
    'slantDashDot',
    'thick',
    'thin',
    'none',
    None,
]

# Define valid underline styles
UnderlineStyle = Literal[
    'single', 'double', 'singleAccounting', 'doubleAccounting', 'none'
]


class FontArgs(TypedDict, total=False):
    bold: bool
    italic: bool
    underline: UnderlineStyle | None
    size: int
    color: Color


def format_range(
    filename: str,
    sheet_name: str,
    start_cell: str,
    end_cell: str | None = None,
    bold: bool = False,
    italic: bool = False,
    underline: bool = False,
    font_size: int | None = None,
    font_color: str | None = None,
    bg_color: str | None = None,
    border_style: BorderStyle = None,
    border_color: str | None = None,
    number_format: str | None = None,
    alignment: str | None = None,
    wrap_text: bool = False,
    merge_cells: bool = False,
    protection: dict[str, Any] | None = None,
    conditional_format: dict[str, Any] | None = None,
) -> dict[str, Any]:
    # Validación de celdas
    if not validate_cell_reference(start_cell):
        return {"error": f"Invalid start cell reference: {start_cell}"}
    if end_cell and not validate_cell_reference(end_cell):
        return {"error": f"Invalid end cell reference: {end_cell}"}
    try:
        wb = get_or_create_workbook(filename)
        if sheet_name not in wb.sheetnames:
            return {"error": f"Sheet '{sheet_name}' not found"}
        sheet = wb[sheet_name]
        try:
            start_row, start_col, end_row, end_col = parse_cell_range(
                start_cell, end_cell
            )
        except ValueError as e:
            return {"error": f"Invalid cell range: {str(e)}"}
        if end_row is None:
            end_row = start_row
        if end_col is None:
            end_col = start_col
        # Font configuration with proper typing
        font_args: FontArgs = {
            "bold": bold,
            "italic": italic,
            "underline": "single" if underline else None,
        }
        if font_size is not None:
            font_args["size"] = font_size
        if font_color is not None:
            font_color = (
                font_color if font_color.startswith("FF") else f"FF{font_color}"
            )
            font_args["color"] = Color(rgb=font_color)

        font = Font(**font_args)

        # Fill configuration
        fill = None
        if bg_color is not None:
            bg_color = bg_color if bg_color.startswith("FF") else f"FF{bg_color}"
            fill = PatternFill(
                start_color=Color(rgb=bg_color),
                end_color=Color(rgb=bg_color),
                fill_type="solid",
            )

        # Border configuration
        border = None
        if border_style is not None and border_style != 'none':
            border_color = border_color if border_color else "000000"
            border_color = (
                border_color if border_color.startswith("FF") else f"FF{border_color}"
            )
            side = Side(style=border_style, color=Color(rgb=border_color))
            border = Border(left=side, right=side, top=side, bottom=side)

        # Alignment
        align = None
        if alignment is not None or wrap_text:
            align = Alignment(
                horizontal=alignment, vertical="center", wrap_text=wrap_text
            )

        # Protection
        protect = Protection(**protection) if protection else None

        # Aplicar formato
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = sheet.cell(row=row, column=col)
                cell.font = font
                if fill:
                    cell.fill = fill
                if border:
                    cell.border = border
                if align:
                    cell.alignment = align
                if protect:
                    cell.protection = protect
                if number_format:
                    cell.number_format = number_format
        # Merge
        if merge_cells and end_cell:
            try:
                sheet.merge_cells(f"{start_cell}:{end_cell}")
            except ValueError as e:
                return {"error": f"Failed to merge cells: {str(e)}"}
        # Conditional formatting
        if conditional_format:
            range_str = f"{start_cell}:{end_cell}" if end_cell else start_cell
            rule_type = conditional_format.get("type")
            if not rule_type:
                return {"error": "Conditional format type not specified"}
            params = conditional_format.get("params", {})
            if rule_type == "cell_is" and "fill" in params:
                fill_params = params["fill"]
                if isinstance(fill_params, dict):
                    fill_color = fill_params.get("fgColor", "FFC7CE")
                    fill_color = (
                        fill_color if fill_color.startswith("FF") else f"FF{fill_color}"
                    )
                    params["fill"] = PatternFill(
                        start_color=fill_color, end_color=fill_color, fill_type="solid"
                    )
            try:
                if rule_type == "color_scale":
                    rule = ColorScaleRule(**params)
                elif rule_type == "data_bar":
                    rule = DataBarRule(**params)
                elif rule_type == "icon_set":
                    rule = IconSetRule(**params)
                elif rule_type == "formula":
                    rule = FormulaRule(**params)
                elif rule_type == "cell_is":
                    rule = CellIsRule(**params)
                else:
                    return {"error": f"Invalid conditional format type: {rule_type}"}
                sheet.conditional_formatting.add(range_str, rule)
            except Exception as e:
                return {"error": f"Failed to apply conditional formatting: {str(e)}"}
        wb.save(filename)
        range_str = f"{start_cell}:{end_cell}" if end_cell else start_cell
        return {
            "message": f"Applied formatting to range {range_str}",
            "range": range_str,
        }
    except Exception as e:
        return {"error": str(e)}
