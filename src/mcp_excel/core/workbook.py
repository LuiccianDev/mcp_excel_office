"""Core functionality for Excel workbook operations.

This module provides functions to create, modify, and inspect Excel workbooks
following the Model Context Protocol (MCP) standards.
"""

from pathlib import Path
from typing import Any, TypedDict

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from .exceptions import SheetExistsError, ValidationError, WorkbookError, WorksheetError

# Type aliases
WorkbookResult = dict[str, Any]  # Type alias for workbook operation results
SheetName = str  # Type alias for Excel sheet names


class WorkbookInfo(TypedDict, total=False):
    """Type definition for workbook information dictionary.

    Attributes:
        filename: Name of the Excel file.
        sheets: List of sheet names in the workbook.
        size: Size of the file in bytes.
        modified: Timestamp of last modification (seconds since epoch).
        used_ranges: Optional dictionary mapping sheet names to their used ranges.
    """

    filename: str
    sheets: list[SheetName]
    size: int
    modified: float
    used_ranges: dict[SheetName, str] | None


def _validate_sheet_name(sheet_name: str) -> None:
    """Validate that the sheet name is valid according to Excel's rules.

    Args:
        sheet_name: Name of the sheet to validate.

    Raises:
        ValidationError: If the sheet name is invalid.
            Possible reasons:
            - Empty or None value
            - Not a string
            - Exceeds 31 characters
            - Contains invalid characters ([]:*?/\\')
            - Starts with a single quote
    """
    if not sheet_name or not isinstance(sheet_name, str):
        raise ValidationError("Sheet name must be a non-empty string")
    if len(sheet_name) > 31:
        raise ValidationError("Sheet name cannot exceed 31 characters")
    if any(char in sheet_name for char in '[]:*?/\\'):
        raise ValidationError(
            'Sheet name cannot contain any of these characters: []:*?/\\'
        )
    if sheet_name.startswith("'"):
        raise ValidationError("Sheet name cannot start with a single quote")


def _create_initial_worksheet(workbook: Workbook, sheet_name: str) -> None:
    """Create the initial worksheet with the specified name.

    This function removes the default worksheet that comes with a new workbook
    and creates a new one with the specified name.

    Args:
        workbook: The workbook to add the sheet to.
        sheet_name: Name for the new worksheet.
    """
    try:
        # Remove default sheet and create a new one with the specified name
        workbook.remove(workbook.active)
        workbook.create_sheet(sheet_name)
    except Exception as e:
        raise WorksheetError(f"Failed to create initial worksheet: {e}") from e


def create_workbook(
    filename: str | Path, sheet_name: str = "Sheet1", data_only: bool = False
) -> WorkbookResult:
    """Create a new Excel workbook with an optional custom sheet name.

    Args:
        filename: Path where the workbook will be saved.
        sheet_name: Name for the initial worksheet. Defaults to "Sheet1".
        data_only: Whether to save only values, not formulas. Defaults to False.

    Returns:
        Dictionary with operation status and details:
        {
            "message": str,         # Success/error message
            "active_sheet": str,   # Name of the active sheet
            "filepath": str        # Full path to the created file
        }
    """
    path = Path(filename).resolve()
    wb = None

    try:
        # Input validation
        _validate_sheet_name(sheet_name)

        # Create workbook and initial worksheet
        wb = Workbook()
        _create_initial_worksheet(wb, sheet_name)

        # Save the workbook
        wb.save(str(path), data_only=data_only)

        return {
            "message": f"Created workbook: {path}",
            "active_sheet": sheet_name,
            "filepath": str(path),
        }

    except PermissionError as e:
        raise PermissionError(f"Cannot write to {path}: {e}") from e
    except Exception as e:
        raise WorkbookError(f"Failed to create workbook: {e}") from e
    finally:
        if wb is not None:
            wb.close()


def _load_existing_workbook(filepath: Path, read_only: bool = False) -> Workbook:
    """Load an existing Excel workbook from the specified path.

    This is a helper function that wraps openpyxl's load_workbook with
    consistent error handling and type hints.

    Args:
        filepath: Path to the Excel file to load.
        read_only: Whether to open the workbook in read-only mode.
            This is more memory-efficient for large files. Defaults to False.

    Returns:
        Workbook: An openpyxl Workbook object.
    """
    try:
        return load_workbook(str(filepath), read_only=read_only, data_only=True)
    except PermissionError as e:
        raise PermissionError(f"Cannot access {filepath}: {e}") from e
    except Exception as e:
        raise WorkbookError(f"Error loading workbook {filepath}: {e}") from e


# * Get or create workbook
def get_or_create_workbook(filename: str | Path, read_only: bool = False) -> Workbook:
    """Get an existing workbook or create a new one if it doesn't exist.

    This is a convenience function that combines loading an existing workbook
    with creating a new one if it doesn't exist, with consistent error handling.

    Args:
        filename: Path to the Excel file. Can be a string or Path object.
        read_only: If True, opens the workbook in read-only mode.
            If the file doesn't exist and read_only is True, raises FileNotFoundError.
            Defaults to False.

    Returns:
        Workbook: An openpyxl Workbook object.
    """
    path = Path(filename).resolve()

    # If file exists, load it
    if path.exists():
        return _load_existing_workbook(path, read_only)

    # If file doesn't exist and read_only is True, raise error
    if read_only:
        raise FileNotFoundError(f"Workbook not found: {path}")

    # Create a new workbook and return it
    wb = Workbook()
    try:
        # Remove the default sheet and create one with a standard name
        _create_initial_worksheet(wb, "Sheet1")
        return wb
    except Exception as e:
        wb.close()
        raise WorkbookError(f"Failed to create new workbook: {e}") from e


def _create_new_sheet(workbook: Workbook, sheet_name: str) -> None:
    """Create a new sheet in the specified workbook.

    This is a helper function that encapsulates the logic for creating a new sheet
    with proper error handling and validation.

    Args:
        workbook: The openpyxl Workbook instance to add the sheet to.
        sheet_name: Name for the new sheet. Will be validated before creation.
    """
    if not isinstance(workbook, Workbook):
        raise TypeError(f"Expected Workbook instance, got {type(workbook).__name__}")

    if sheet_name in workbook.sheetnames:
        raise SheetExistsError(f"Sheet '{sheet_name}' already exists")

    try:
        workbook.create_sheet(sheet_name)
    except ValueError as e:
        raise ValidationError(f"Invalid sheet name '{sheet_name}': {e}") from e
    except Exception as e:
        raise WorksheetError(f"Failed to create sheet '{sheet_name}': {e}") from e


# * Create sheet
def create_sheet(filename: str | Path, sheet_name: str) -> WorkbookResult:
    """Create a new sheet in the specified Excel workbook.

    If the specified file doesn't exist, a new workbook will be created with the sheet.
    If the file exists, the new sheet will be added to the existing workbook.

    Args:
        filename: Path to the Excel file. Can be a string or Path object.
        sheet_name: Name for the new sheet. Will be validated before creation.

    Returns:
        Dictionary with operation status and details:
        {
            "message": str,         # Success/error message
            "sheet_name": str,      # Name of the created sheet
            "filepath": str         # Full path to the file
        }
    """
    path = Path(filename).resolve()

    # If file doesn't exist, create a new workbook with the sheet
    if not path.exists():
        return create_workbook(path, sheet_name)

    wb = None
    try:
        # Load existing workbook
        wb = _load_existing_workbook(path)

        # Validate sheet name and create the sheet
        _validate_sheet_name(sheet_name)
        _create_new_sheet(wb, sheet_name)

        # Save the workbook
        wb.save(str(path))

        return {
            "message": f"Created sheet '{sheet_name}' in {path}",
            "sheet_name": sheet_name,
            "filepath": str(path),
        }

    except (SheetExistsError, ValidationError, PermissionError):
        # Re-raise these specific exceptions as-is
        raise
    except Exception as e:
        # Wrap other exceptions in WorkbookError
        raise WorkbookError(f"Failed to create sheet '{sheet_name}': {e}") from e
    finally:
        # Ensure the workbook is properly closed
        if wb is not None:
            wb.close()


def _get_worksheet_range(worksheet: Any) -> str | None:
    """Get the used range of a worksheet as a string (e.g., 'A1:D10').

    Args:
        worksheet: The openpyxl worksheet to analyze.

    Returns:
        A string representing the used range (e.g., 'A1:D10') or None if the
        worksheet is empty.
    """
    if worksheet.max_row > 0 and worksheet.max_column > 0:
        return f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
    return None


# * Get workbook info
def get_workbook_info(
    filename: str | Path, include_ranges: bool = False
) -> WorkbookInfo:
    """Get metadata about an Excel workbook including sheets, file info, and ranges.

    This function provides comprehensive information about an Excel workbook,
    including its sheets, file metadata, and optionally the used ranges for
    each worksheet.

    Args:
        filename: Path to the Excel file. Can be a string or Path object.
        include_ranges: If True, includes used ranges for each sheet.
            Note: This will load the entire workbook to calculate ranges.
            Defaults to False for better performance with large files.

    Returns:
        A dictionary containing the following keys:
        - filename: str - Name of the file
        - sheets: List[str] - Names of all worksheets in the workbook
        - size: int - File size in bytes
        - modified: float - Last modification timestamp (seconds since epoch)
        - used_ranges: Optional[Dict[str, str]] - If include_ranges is True,
            a dictionary mapping sheet names to their used ranges (e.g., 'A1:D10')
    """
    path = Path(filename).resolve()

    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")

    wb = None
    try:
        # Load in read-only mode for better performance with large files
        wb = load_workbook(str(path), read_only=True, data_only=True)

        # Basic file information
        info: WorkbookInfo = {
            "filename": path.name,
            "sheets": wb.sheetnames,
            "size": path.stat().st_size,
            "modified": path.stat().st_mtime,
            "used_ranges": None,
        }

        # Calculate used ranges if requested
        if include_ranges:
            ranges: dict[str, str] = {}
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                if range_str := _get_worksheet_range(ws):
                    ranges[sheet_name] = range_str

            info["used_ranges"] = ranges or None

        return info

    except PermissionError as e:
        raise PermissionError(f"Cannot access {path}: {e}") from e
    except Exception as e:
        raise WorkbookError(f"Failed to get workbook info: {e}") from e
    finally:
        if wb is not None:
            wb.close()
