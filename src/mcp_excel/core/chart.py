from enum import Enum
from typing import Any, cast

from openpyxl import load_workbook
from openpyxl.chart import (
    AreaChart,
    BarChart,
    LineChart,
    PieChart,
    Reference,
    ScatterChart,
)
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.legend import Legend
from openpyxl.worksheet.worksheet import Worksheet

from mcp_excel.utils.cell_utils import parse_cell_range


class ChartType(str, Enum):
    """Supported chart types"""

    LINE = "line"
    BAR = "bar"
    PIE = "pie"
    SCATTER = "scatter"
    AREA = "area"
    BUBBLE = "bubble"
    STOCK = "stock"
    SURFACE = "surface"
    RADAR = "radar"


class ChartStyle:
    """Chart style configuration"""

    def __init__(
        self,
        title_size: int = 14,
        title_bold: bool = True,
        axis_label_size: int = 12,
        show_legend: bool = True,
        legend_position: str = "r",
        show_data_labels: bool = True,
        grid_lines: bool = False,
        style_id: int = 2,
    ):
        self.title_size = title_size
        self.title_bold = title_bold
        self.axis_label_size = axis_label_size
        self.show_legend = show_legend
        self.legend_position = legend_position
        self.show_data_labels = show_data_labels
        self.grid_lines = grid_lines
        self.style_id = style_id


def create_chart_in_sheet(
    filename: str,
    sheet_name: str,
    data_range: str,
    chart_type: str,
    target_cell: str,
    title: str = "",
    x_axis: str = "",
    y_axis: str = "",
    style: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Create chart in sheet with enhanced styling options"""
    try:
        wb = load_workbook(filename)
        if sheet_name not in wb.sheetnames:
            return {"status": "error", "message": f"Sheet '{sheet_name}' not found"}
        worksheet = cast(Worksheet, wb[sheet_name])
        if "!" in data_range:
            range_sheet_name, cell_range = data_range.split("!")
            if range_sheet_name not in wb.sheetnames:
                return {
                    "status": "error",
                    "message": f"Sheet '{range_sheet_name}' referenced in data range not found",
                }
            worksheet = cast(Worksheet, wb[range_sheet_name])
        else:
            cell_range = data_range
        try:
            start_cell, end_cell = cell_range.split(":")
            start_row, start_col, end_row, end_col = parse_cell_range(
                start_cell, end_cell
            )
            if None in (start_row, start_col, end_row, end_col):
                return {
                    "status": "error",
                    "message": "Failed to parse cell range",
                }
            assert start_row is not None
            assert start_col is not None
            assert end_row is not None
            assert end_col is not None
        except ValueError as e:
            return {
                "status": "error",
                "message": f"Invalid data range format: {str(e)}",
            }
        chart_classes = {
            "line": LineChart,
            "bar": BarChart,
            "pie": PieChart,
            "scatter": ScatterChart,
            "area": AreaChart,
        }
        chart_type_lower = chart_type.lower()
        chart_class = chart_classes.get(chart_type_lower)
        if not chart_class:
            return {
                "status": "error",
                "message": (
                    f"Unsupported chart type: {chart_type}. "
                    f"Supported types: {', '.join(chart_classes.keys())}"
                ),
            }
        chart = chart_class()
        chart.title = title
        if hasattr(chart, "x_axis"):
            chart.x_axis.title = x_axis
        if hasattr(chart, "y_axis"):
            chart.y_axis.title = y_axis
        try:
            if chart_type_lower == "scatter":
                for col in range(start_col + 1, end_col + 1):
                    x_values = Reference(
                        worksheet,
                        min_row=start_row + 1,
                        max_row=end_row,
                        min_col=start_col,
                    )
                    y_values = Reference(
                        worksheet, min_row=start_row + 1, max_row=end_row, min_col=col
                    )
                    chart.series.append((x_values, y_values))  # simplified for mypy
            else:
                data = Reference(
                    worksheet,
                    min_row=start_row,
                    max_row=end_row,
                    min_col=start_col + 1,
                    max_col=end_col,
                )
                cats = Reference(
                    worksheet, min_row=start_row + 1, max_row=end_row, min_col=start_col
                )
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
        except Exception as e:
            return {
                "status": "error",
                "message": f"Failed to create chart data references: {str(e)}",
            }
        try:
            if style:
                if style.get("show_legend", True):
                    chart.legend = Legend()
                    chart.legend.position = style.get("legend_position", "r")
                else:
                    chart.legend = None
                if style.get("show_data_labels", False):
                    chart.dataLabels = DataLabelList()
                    chart.dataLabels.showVal = True
                if style.get("grid_lines", False):
                    if hasattr(chart, "x_axis"):
                        chart.x_axis.majorGridlines = ChartLines()
                    if hasattr(chart, "y_axis"):
                        chart.y_axis.majorGridlines = ChartLines()
        except Exception as e:
            return {
                "status": "error",
                "message": f"Failed to apply chart style: {str(e)}",
            }
        chart.width = 15
        chart.height = 7.5
        try:
            worksheet.add_chart(chart, target_cell)
        except ValueError as e:
            return {"status": "error", "message": f"Invalid target cell: {str(e)}"}
        except Exception as e:
            return {
                "status": "error",
                "message": f"Failed to create chart drawing: {str(e)}",
            }
        try:
            wb.save(filename)
        except Exception as e:
            return {
                "status": "error",
                "message": f"Failed to save workbook with chart: {str(e)}",
            }
        return {
            "status": "success",
            "message": f"{chart_type.capitalize()} chart created successfully",
            "details": {
                "type": chart_type,
                "location": target_cell,
                "data_range": data_range,
            },
        }
    except Exception as e:
        return {
            "status": "error",
            "message": f"Unexpected error creating chart: {str(e)}",
        }
