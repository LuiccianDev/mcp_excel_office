"""Core functionality for reading and writing Excel data.

This module provides functions to interact with Excel files, including reading ranges,
writing data, and detecting headers, following the Model Context Protocol (MCP) standards.
"""

from pathlib import Path
from typing import Any, TypedDict

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from mcp_excel.utils.cell_utils import parse_cell_range

from .exceptions import (
    DataError,
    InvalidCellReferenceError,
    InvalidDataError,
    RangeError,
    SheetError,
    SheetNotFoundError,
    WorkbookError,
)

# Type aliases for better code readability
CellValue = Any
RowData = list[CellValue]
SheetData = list[RowData]
CellRange = str


class RangeCoordinates(TypedDict):
    """Type definition for cell range coordinates."""

    start_row: int
    start_col: int
    end_row: int
    end_col: int


def _get_worksheet(workbook: Any, sheet_name: str) -> Worksheet:
    """Get a worksheet by name, raising appropriate exceptions.

    Args:
        workbook: Openpyxl workbook object.
        sheet_name: Name of the worksheet to retrieve.

    Returns:
        The requested worksheet.

    Raises:
        SheetNotFoundError: If the specified sheet doesn't exist.
    """
    if sheet_name not in workbook.sheetnames:
        raise SheetNotFoundError(f"Sheet '{sheet_name}' not found")
    return workbook[sheet_name]


def _parse_cell_reference(cell_ref: str) -> tuple[int, int]:
    """Parse cell reference into (row, column) coordinates.

    Args:
        cell_ref: Cell reference (e.g., 'A1', 'B2').

    Returns:
        Tuple of (row, column) coordinates (1-based).

    Raises:
        InvalidCellReferenceError: If the cell reference is invalid.
    """
    try:
        coords = parse_cell_range(f"{cell_ref}:{cell_ref}")
        if not coords or None in coords[:2]:
            raise InvalidCellReferenceError(f"Invalid cell reference: {cell_ref}")
        return coords[0], coords[1]  # row, column
    except ValueError as e:
        raise InvalidCellReferenceError(f"Invalid cell format: {str(e)}") from e


def _get_used_range(
    worksheet: Worksheet, start_row: int, start_col: int
) -> tuple[int, int]:
    """Find the used range in a worksheet starting from given coordinates.

    Args:
        worksheet: The worksheet to search.
        start_row: Starting row (1-based).
        start_col: Starting column (1-based).

    Returns:
        Tuple of (end_row, end_col) coordinates.
    """
    end_row = start_row
    end_col = start_col

    # Find last used row
    while end_row <= worksheet.max_row and any(
        worksheet.cell(row=end_row, column=c).value is not None
        for c in range(start_col, worksheet.max_column + 1)
    ):
        end_row += 1

    # Find last used column
    while end_col <= worksheet.max_column and any(
        worksheet.cell(row=r, column=end_col).value is not None
        for r in range(start_row, worksheet.max_row + 1)
    ):
        end_col += 1

    return end_row - 1, end_col - 1  # Convert to 0-based exclusive


# * Read a range of data from an Excel worksheet
def read_excel_range(
    filename: Path | str,
    sheet_name: str,
    start_cell: str = "A1",
    end_cell: str | None = None,
    preview_only: bool = False,
) -> SheetData:
    """Read a range of data from an Excel worksheet.

    Args:
        filename: Path to the Excel file.
        sheet_name: Name of the worksheet to read from.
        start_cell: Starting cell reference (e.g., 'A1').
        end_cell: Optional ending cell reference. If not provided, will detect used range.
        preview_only: If True, only returns a preview of the data.

    Returns:
        A 2D list containing the cell values in the specified range.

    Raises:
        FileNotFoundError: If the specified file doesn't exist.
        SheetNotFoundError: If the specified sheet doesn't exist.
        InvalidCellReferenceError: If cell references are invalid.
        DataError: For errors related to data processing.
    """
    path = Path(filename).resolve()
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")

    wb = None
    try:
        wb = load_workbook(str(path), read_only=True, data_only=True)
        ws = _get_worksheet(wb, sheet_name)

        # Handle range string (e.g., "A1:B2")
        if ":" in start_cell:
            start_cell, end_cell = start_cell.split(":", 1)

        # Parse start cell coordinates
        start_row, start_col = _parse_cell_reference(start_cell)

        # Parse or detect end cell coordinates
        if end_cell:
            end_row, end_col = _parse_cell_reference(end_cell)
        else:
            end_row, end_col = _get_used_range(ws, start_row, start_col)

        # Validate range is within worksheet bounds
        if start_row > ws.max_row or start_col > ws.max_column:
            max_cell = f"{get_column_letter(ws.max_column)}{ws.max_row}"
            raise RangeError(
                f"Start cell out of bounds. Sheet dimensions are A1:{max_cell}"
            )

        # Read the data
        data: SheetData = []
        for row in range(start_row, min(end_row, ws.max_row) + 1):
            row_data = [
                ws.cell(row=row, column=col).value
                for col in range(start_col, min(end_col, ws.max_column) + 1)
            ]
            if any(v is not None for v in row_data):
                data.append(row_data)

        return data

    except Exception as e:
        if not isinstance(
            e,
            FileNotFoundError
            | SheetNotFoundError
            | InvalidCellReferenceError
            | DataError,
        ):
            raise DataError(f"Failed to read Excel range: {str(e)}") from e
        raise
    finally:
        if wb is not None:
            wb.close()


def _write_data_to_worksheet(
    worksheet: Worksheet, data: SheetData, start_row: int, start_col: int
) -> None:
    """Write data to a worksheet starting at the specified coordinates.

    Args:
        worksheet: The worksheet to write to.
        data: 2D list of data to write.
        start_row: Starting row (1-based).
        start_col: Starting column (1-based).

    Note:
        None values are skipped to preserve existing cell formatting.
    """
    for row_idx, row_data in enumerate(data):
        for col_idx, value in enumerate(row_data):
            if value is not None:  # Skip None values to preserve cell formatting
                worksheet.cell(
                    row=start_row + row_idx, column=start_col + col_idx, value=value
                )


# * Write data to an Excel worksheet
def write_data(
    filename: str | Path,
    sheet_name: str | None,
    data: SheetData | None,
    start_cell: str = "A1",
) -> dict[str, str]:
    """Write data to an Excel worksheet.

    Args:
        filename: Path to the Excel file.
        sheet_name: Name of the worksheet to write to. If None, uses active sheet.
        data: 2D list of data to write.
        start_cell: Starting cell reference (e.g., 'A1').

    Returns:
        Dictionary with operation status and details.

    Raises:
        InvalidDataError: If input data is invalid.
        InvalidCellReferenceError: If cell reference is invalid.
        SheetError: For sheet-related errors.
        WorkbookError: For other workbook-related errors.
    """
    if not data:
        raise InvalidDataError("No data provided to write")

    path = Path(filename).resolve()
    wb = None
    try:
        # Load or create workbook
        try:
            if path.exists():
                wb = load_workbook(str(path))
            else:
                wb = Workbook()
                # Remove default sheet if it exists
                if wb.sheetnames:
                    wb.remove(wb.active)
        except Exception as e:
            raise WorkbookError(f"Failed to access workbook: {str(e)}") from e

        # Get or create worksheet
        try:
            if not sheet_name:
                if not wb.sheetnames:  # If no sheets exist
                    ws = wb.create_sheet("Sheet1")
                else:
                    ws = wb.active
            else:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                else:
                    ws = wb.create_sheet(sheet_name)
        except Exception as e:
            raise SheetError(f"Failed to access worksheet: {str(e)}") from e

        # Parse start cell
        try:
            start_row, start_col = _parse_cell_reference(start_cell)
        except InvalidCellReferenceError as e:
            raise InvalidCellReferenceError(
                f"Invalid start cell format: {str(e)}"
            ) from e

        # Ensure parent directory exists
        path.parent.mkdir(parents=True, exist_ok=True)

        # Write data
        _write_data_to_worksheet(ws, data, start_row, start_col)

        # Save changes
        try:
            wb.save(str(path))
        except Exception as e:
            raise WorkbookError(f"Failed to save workbook: {str(e)}") from e

        return {
            "message": f"Data written to {ws.title}",
            "active_sheet": ws.title,
            "filepath": str(path),
        }

    except Exception as e:
        if not isinstance(
            e,
            InvalidDataError | InvalidCellReferenceError | SheetError | WorkbookError,
        ):
            raise WorkbookError(f"Failed to write data: {str(e)}") from e
        raise
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass  # noqa: S110
