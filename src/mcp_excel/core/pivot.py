import uuid
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from mcp_excel.core.data import read_excel_range
from mcp_excel.exceptions.exceptions import PivotError, ValidationError
from mcp_excel.utils.cell import parse_cell_range


def create_pivot_table(
    filename: str,
    sheet_name: str,
    data_range: str,
    rows: list[str],
    values: list[str],
    columns: list[str] | None = None,
    agg_func: str = "sum",
) -> dict[str, Any]:
    """Create pivot table in sheet using Excel table functionality"""
    try:
        wb = load_workbook(filename)
        if sheet_name not in wb.sheetnames:
            raise ValidationError(f"Sheet '{sheet_name}' not found")
        if ":" not in data_range:
            raise ValidationError("Data range must be in format 'A1:B2'")
        try:
            start_cell, end_cell = data_range.split(":")
            start_row, start_col, end_row, end_col = parse_cell_range(
                start_cell, end_cell
            )
        except ValueError as e:
            raise ValidationError(f"Invalid data range format: {str(e)}")
        if end_row is None or end_col is None:
            raise ValidationError("Invalid data range format: missing end coordinates")
        data_range_str = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
        try:
            data = read_excel_range(filename, sheet_name, start_cell, end_cell)
            if not data:
                raise PivotError("No data found in range")
        except Exception as e:
            raise PivotError(f"Failed to read source data: {str(e)}")
        valid_agg_funcs = ["sum", "average", "count", "min", "max"]
        if agg_func.lower() not in valid_agg_funcs:
            raise ValidationError(
                f"Invalid aggregation function. Must be one of: {', '.join(valid_agg_funcs)}"
            )

        def clean_field_name(field: str) -> str:
            field = str(field).strip()
            for suffix in [" (sum)", " (average)", " (count)", " (min)", " (max)"]:
                if field.lower().endswith(suffix):
                    return field[: -len(suffix)]
            return field

        if data:
            first_row = data[0]
            available_fields = {
                clean_field_name(str(header)).lower() for header in first_row.keys()
            }
            for field_list, field_type in [(rows, "row"), (values, "value")]:
                for field in field_list:
                    if clean_field_name(str(field)).lower() not in available_fields:
                        raise ValidationError(
                            f"Invalid {field_type} field '{field}'. "
                            f"Available fields: {', '.join(sorted(available_fields))}"
                        )
            if columns:
                for field in columns:
                    if clean_field_name(str(field)).lower() not in available_fields:
                        raise ValidationError(
                            f"Invalid column field '{field}'. "
                            f"Available fields: {', '.join(sorted(available_fields))}"
                        )
            if all(
                any(
                    clean_field_name(str(header)).lower()
                    == clean_field_name(str(field)).lower()
                    for field in rows + values
                )
                for header in first_row.keys()
            ):
                data = data[1:]
        cleaned_rows = [clean_field_name(field) for field in rows]
        cleaned_values = [clean_field_name(field) for field in values]
        cleaned_columns = (
            [clean_field_name(field) for field in columns] if columns else []
        )
        pivot_sheet_name = f"{sheet_name}_pivot"
        if pivot_sheet_name in wb.sheetnames:
            wb.remove(wb[pivot_sheet_name])
        pivot_ws = wb.create_sheet(pivot_sheet_name)
        # Write headers
        current_row = 1
        current_col = 1
        for field in cleaned_rows:
            cell = pivot_ws.cell(row=current_row, column=current_col, value=field)
            cell.font = Font(bold=True)
            current_col += 1
        # If columns, add column headers
        col_headers = []
        if cleaned_columns:
            unique_col_combinations = _get_combinations(
                {
                    col: sorted({str(record.get(col, "")) for record in data})
                    for col in cleaned_columns
                }
            )
            for combo in unique_col_combinations:
                for value_field in cleaned_values:
                    header = (
                        " | ".join([str(combo[c]) for c in cleaned_columns])
                        + f" | {value_field} ({agg_func})"
                    )
                    col_headers.append((combo, value_field, header))
                    cell = pivot_ws.cell(
                        row=current_row, column=current_col, value=header
                    )
                    cell.font = Font(bold=True)
                    current_col += 1
        else:
            for value_field in cleaned_values:
                cell = pivot_ws.cell(
                    row=current_row,
                    column=current_col,
                    value=f"{value_field} ({agg_func})",
                )
                cell.font = Font(bold=True)
                current_col += 1
        # Get unique values for each row field
        field_values = {}
        for field in cleaned_rows:
            all_values = []
            for record in data:
                value = str(record.get(field, ""))
                all_values.append(value)
            field_values[field] = sorted(set(all_values))
        row_combinations = _get_combinations(field_values)
        total_rows = len(row_combinations) + 1
        total_cols = len(cleaned_rows) + (
            len(col_headers) if col_headers else len(cleaned_values)
        )
        # Write data rows
        current_row = 2
        for combo in row_combinations:
            col = 1
            for field in cleaned_rows:
                pivot_ws.cell(row=current_row, column=col, value=combo[field])
                col += 1
            filtered_data = _filter_data(data, combo, {})
            if cleaned_columns:
                for col_combo, value_field, _ in col_headers:
                    filtered = _filter_data(filtered_data, {}, col_combo)
                    value = _aggregate_values(filtered, value_field, agg_func)
                    pivot_ws.cell(row=current_row, column=col, value=value)
                    col += 1
            else:
                for value_field in cleaned_values:
                    value = _aggregate_values(filtered_data, value_field, agg_func)
                    pivot_ws.cell(row=current_row, column=col, value=value)
                    col += 1
            current_row += 1
        # Create a table for the pivot data
        try:
            pivot_range = f"A1:{get_column_letter(total_cols)}{total_rows}"
            pivot_table = Table(
                displayName=f"PivotTable_{uuid.uuid4().hex[:8]}", ref=pivot_range
            )
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=True,
            )
            pivot_table.tableStyleInfo = style
            pivot_ws.add_table(pivot_table)
        except Exception as e:
            raise PivotError(f"Failed to create pivot table formatting: {str(e)}")
        try:
            wb.save(filename)
        except Exception as e:
            raise PivotError(f"Failed to save workbook: {str(e)}")
        return {
            "message": "Summary table created successfully",
            "details": {
                "source_range": data_range_str,
                "pivot_sheet": pivot_sheet_name,
                "rows": cleaned_rows,
                "columns": cleaned_columns,
                "values": cleaned_values,
                "aggregation": agg_func,
            },
        }
    except (ValidationError, PivotError) as e:
        return {"error": str(e), "message": "Failed to create pivot table"}
    except Exception as e:
        raise PivotError(str(e))


def _get_combinations(field_values: dict[str, set]) -> list[dict]:
    """Get all combinations of field values."""
    result = [{}]
    for field, values in list(
        field_values.items()
    ):  # Convert to list to avoid runtime changes
        new_result = []
        for combo in result:
            for value in sorted(values):  # Sort for consistent ordering
                new_combo = combo.copy()
                new_combo[field] = value
                new_result.append(new_combo)
        result = new_result
    return result


def _filter_data(data: list[dict], row_filters: dict, col_filters: dict) -> list[dict]:
    """Filter data based on row and column filters."""
    result = []
    for record in data:
        matches = True
        for field, value in row_filters.items():
            if record.get(field) != value:
                matches = False
                break
        for field, value in col_filters.items():
            if record.get(field) != value:
                matches = False
                break
        if matches:
            result.append(record)
    return result


def _aggregate_values(data: list[dict], field: str, agg_func: str) -> float:
    """Aggregate values using the specified function."""
    values = [
        record[field]
        for record in data
        if field in record and isinstance(record[field], (int, float))
    ]
    if not values:
        return 0

    if agg_func == "sum":
        return sum(values)
    elif agg_func == "average":
        return sum(values) / len(values)
    elif agg_func == "count":
        return len(values)
    elif agg_func == "min":
        return min(values)
    elif agg_func == "max":
        return max(values)
    else:
        return sum(values)  # Default to sum
