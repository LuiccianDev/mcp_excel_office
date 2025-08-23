import re
from collections.abc import Generator
from contextlib import contextmanager
from typing import Any

import psycopg2
from openpyxl import load_workbook
from psycopg2 import sql
from psycopg2.extensions import connection
from psycopg2.extras import DictCursor

from mcp_excel.utils.db_config_utils import settings


class DatabaseError(Exception):
    """Custom exception for database-related errors."""

    pass


# * Fetch data from database with proper connection handling
def fetch_data_from_db(
    connection_string: str | None = None, query: str = "", params: tuple | None = None
) -> dict[str, Any]:
    """Fetch data from database with proper connection handling.

    Args:
        connection_string: Optional database connection string. If not provided,
                         uses the default connection from settings.
        query: SQL query to execute
        params: Optional query parameters

    Returns:
        Dictionary with 'columns' and 'rows' keys, or 'error' key if failed
    """
    if not query:
        return {"status": "error", "message": "No query provided"}

    conn_str = connection_string or settings.database_uri
    """Fetch data from database with proper connection handling.

    Args:
        connection_string: Database connection string
        query: SQL query to execute
        params: Optional query parameters

    Returns:
        Dictionary with 'columns' and 'rows' keys, or 'error' key if failed
    """
    try:
        with _get_db_connection(conn_str) as conn:
            rows, columns = _execute_query(conn, query, params)
            return {
                "status": "success",
                "columns": columns,
                "rows": [tuple(row.values()) for row in rows],
            }
    except DatabaseError as e:
        return {"status": "error", "message": str(e)}


# * Insert data into Excel file
def insert_data_to_excel(
    filename: str, sheet_name: str, columns: list, rows: list
) -> dict[str, Any]:
    try:
        wb = load_workbook(filename)
        ws = wb[sheet_name]
        # Escribir encabezados
        for col_num, col_name in enumerate(columns, 1):
            ws.cell(row=1, column=col_num, value=col_name)
        # Escribir datos
        for row_num, row_data in enumerate(rows, 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        wb.save(filename)
        return {
            "status": "success",
            "message": f"Inserted {len(rows)} rows into '{sheet_name}'",
        }
    except Exception as e:
        return {"status": "error", "message": str(e)}


# * Validate SQL query to prevent SQL injection
def validate_sql_query(query: str) -> bool:
    """
    Validate SQL query to prevent SQL injection.

    Args:
        query: SQL query to validate

    Returns:
        bool: True if query is safe, False otherwise
    """
    # Normalize whitespace and convert to lowercase for easier checking
    normalized = " ".join(query.lower().split())

    # Check for multiple statements (prevent SQL injection)
    if ";" in normalized.replace(";", " ; ").split():
        return False

    # Check for dangerous SQL commands
    dangerous_commands = [
        "drop",
        "delete",
        "insert",
        "update",
        "alter",
        "truncate",
        "create",
        "modify",
        "grant",
        "revoke",
        "exec",
        "execute",
        "shutdown",
        "--",
        "/*",
        "*/",
        "xp_",
        "sp_",
    ]

    if any(cmd in normalized for cmd in dangerous_commands):
        return False

    # Check if it's a SELECT query
    if not normalized.startswith("select "):
        return False

    # Additional validation for table names
    if not re.match(r'^select\s+[\w\s,*]+\s+from\s+[\w"]+', normalized):
        return False

    return True


# * Clean data rows: remove None, strip strings, and basic type normalization.
def clean_data(rows: list, columns: list) -> list:
    """
    Clean data rows: remove None, strip strings, and basic type normalization.
    Validates that each row matches the expected number of columns.
    """
    cleaned = []
    expected_len = len(columns)

    for i, row in enumerate(rows):
        if len(row) != expected_len:
            raise ValueError(f"Row {i} has {len(row)} values, expected {expected_len}.")

        cleaned_row = []
        for value in row:
            if isinstance(value, str):
                cleaned_row.append(value.strip())
            elif value is None:
                cleaned_row.append("")
            else:
                cleaned_row.append(value)
        cleaned.append(tuple(cleaned_row))

    return cleaned


# * Insert data into the database with batch processing.
def insert_data_to_db(
    table: str,
    columns: list[str],
    rows: list[tuple],
    connection_string: str | None = None,
    batch_size: int = 1000,
) -> dict[str, Any]:
    """
    Insert data into the database with batch processing.

    Args:
        connection_string: Database connection string
        table: Target table name
        columns: List of column names
        rows: List of tuples with data to insert
        batch_size: Number of rows to insert in each batch

    Returns:
        Dictionary with operation result or error message
    """
    if not all(isinstance(col, str) for col in columns):
        return {"status": "error", "message": "All column names must be strings"}

    if not all(len(row) == len(columns) for row in rows):
        return {"status": "error", "message": "Row length must match number of columns"}

    try:
        conn_str = connection_string or settings.database_uri
        with _get_db_connection(conn_str) as conn, conn.cursor() as cursor:
            # Use sql.SQL and sql.Identifier for safe SQL composition
            query = sql.SQL(
                """
                INSERT INTO {table} ({fields})
                VALUES %s
                ON CONFLICT DO NOTHING
            """
            ).format(
                table=sql.Identifier(table),
                fields=sql.SQL(", ").join(map(sql.Identifier, columns)),
            )

            # Process in batches
            total_inserted = 0
            for i in range(0, len(rows), batch_size):
                batch = rows[i : i + batch_size]
                try:
                    psycopg2.extras.execute_values(
                        cursor, query, batch, template=None, page_size=batch_size
                    )
                    total_inserted += cursor.rowcount
                    conn.commit()
                except psycopg2.Error as e:
                    conn.rollback()

                    raise DatabaseError(f"Batch insert failed: {e}") from e

            return {
                "status": "success",
                "message": f"Successfully inserted {total_inserted} rows into '{table}'",
            }
    except DatabaseError as e:
        return {"status": "error", "message": str(e)}
    except Exception as e:
        return {"status": "error", "message": f"An unexpected error occurred: {str(e)}"}

# Context manager for database connections
@contextmanager
def _get_db_connection(
    connection_string: str | None = None,
) -> Generator[connection, None, None]:
    """Context manager for database connections.

    Args:
        connection_string: Database connection string.

    Yields:
        connection: psycopg2 connection object with DictCursor.

    Raises:
        DatabaseError: If connection cannot be established.
    """
    conn: connection | None = None
    conn_str = connection_string or settings.database_uri
    try:
        conn = psycopg2.connect(
            conn_str, cursor_factory=DictCursor, connect_timeout=10
        )
        conn.autocommit = False
        yield conn
    except psycopg2.Error as e:
        raise DatabaseError(f"Failed to connect to database: {e}") from e
    finally:
        if conn is not None:
            conn.close()


# Execute a query and return results
def _execute_query(
    conn: connection, query: str, params: tuple | None = None
) -> tuple[list[dict[str, Any]], list[str]]:
    """Execute a query and return results.

    Args:
        conn: Database connection
        query: SQL query to execute
        params: Query parameters

    Returns:
        Tuple of (rows as dictionaries, column names)
    """
    try:
        with conn.cursor() as cursor:
            cursor.execute(query, params or ())
            if cursor.description:
                columns = [desc[0] for desc in cursor.description]
                rows = [
                    dict(zip(columns, row, strict=False)) for row in cursor.fetchall()
                ]
                return rows, columns
            return [], []
    except psycopg2.Error as e:
        conn.rollback()
        raise DatabaseError(f"Query execution failed: {e}") from e
